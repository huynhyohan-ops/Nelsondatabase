from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ====== CONFIG PATHS ======

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "Data"
REPORTS_DIR = BASE_DIR / "Output" / "Reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

SHIPMENTS_FILE = DATA_DIR / "Shipments.xlsx"

# Nếu sau này anh có sheet summary riêng thì gán tên vào đây,
# còn hiện tại layout đang dùng sheet active cho cả I và II.
SUMMARY_SHEET_NAME: Optional[str] = None

# ----- Bảng I / TARGET THEO THÁNG -----
# Tuần: nằm ở cột B, các dòng 7–10
TARGET_FIRST_WEEK_ROW = 7
TARGET_WEEKS_COUNT = 4
TARGET_WEEK_COL = 2          # cột B
TARGET_PROFIT_COL = 7        # cột G
TARGET_VOLUME_COL = 8        # cột H
TARGET_TOTAL_ROW = TARGET_FIRST_WEEK_ROW + TARGET_WEEKS_COUNT  # dòng TOTAL

# ----- II.1 / KHÁCH HÀNG ĐÃ SỬ DỤNG DV TRONG TUẦN -----
DETAIL_HEADER_ROW = 15       # dòng header: STT, ACCOUNT, ROUTING, VOLUME, PROFIT
DETAIL_FIRST_ROW = 16        # dòng data đầu tiên
DETAIL_MAX_ROWS = 200        # tối đa số dòng data để xoá/clear


# ====== ISO WEEK ======

@dataclass(frozen=True)
class IsoWeek:
    year: int
    week: int

    @classmethod
    def from_date(cls, value: date) -> "IsoWeek":
        iso_year, iso_week, _ = value.isocalendar()
        return cls(iso_year, iso_week)

    def add_weeks(self, delta: int) -> "IsoWeek":
        monday = date.fromisocalendar(self.year, self.week, 1)
        monday += timedelta(days=7 * delta)
        iso_year, iso_week, _ = monday.isocalendar()
        return IsoWeek(iso_year, iso_week)

    def label(self) -> str:
        return f"{self.week}/{self.year}"


def iso_week_range(year: int, week: int) -> tuple[date, date]:
    jan4 = date(year, 1, 4)
    jan4_weekday = jan4.isoweekday()
    monday = jan4 + timedelta(days=(week - 1) * 7 - (jan4_weekday - 1))
    sunday = monday + timedelta(days=6)
    return monday, sunday


def get_current_iso_week_and_year() -> tuple[int, int]:
    today = datetime.today().date()
    iso_year, iso_week, _ = today.isocalendar()
    return iso_year, iso_week


# ====== LOAD SHIPMENTS THEO TUẦN ======

def load_shipments_for_week(year: int, week: int) -> List[dict]:
    if not SHIPMENTS_FILE.exists():
        raise FileNotFoundError(f"Không tìm thấy file shipments: {SHIPMENTS_FILE}")

    sheets = pd.read_excel(SHIPMENTS_FILE, sheet_name=None)
    records: list[dict] = []

    for _, df in sheets.items():
        if df is None or df.empty:
            continue

        for col in ["Customer", "Routing", "ETD", "Volume", "Profit"]:
            if col not in df.columns:
                df[col] = None

        df["ETD"] = pd.to_datetime(df["ETD"], dayfirst=True, errors="coerce")

        for _, row in df.iterrows():
            etd = row["ETD"]
            if pd.isna(etd):
                continue

            iso_year, iso_week, _ = etd.isocalendar()
            if iso_year != year or iso_week != week:
                continue

            volume = float(row["Volume"]) if pd.notna(row["Volume"]) else 0.0
            profit = float(row["Profit"]) if pd.notna(row["Profit"]) else 0.0

            records.append(
                {
                    "etd": etd.date(),
                    "customer": str(row["Customer"]) if pd.notna(row["Customer"]) else "",
                    "routing": str(row["Routing"]) if pd.notna(row["Routing"]) else "",
                    "volume": volume,
                    "profit": profit,
                }
            )

    records.sort(key=lambda r: r["etd"])
    return records


# ====== FILE REPORT MANAGEMENT ======

def iter_report_files() -> List[Path]:
    if not REPORTS_DIR.exists():
        return []
    candidates: list[Path] = []
    for path in REPORTS_DIR.glob("*.xlsx"):
        name = path.name.upper()
        if "REPORT" in name and "WEEKLY" in name:
            candidates.append(path)
    return sorted(candidates)


def build_report_filename(week: IsoWeek, user_name: str) -> str:
    safe_user = user_name.strip().replace(" ", "_") or "Unknown"
    return f"Report Weekly {week.week:02d}-{week.year} - {safe_user}.xlsx"


def choose_template_source() -> Path:
    existing = iter_report_files()
    if existing:
        return existing[-1]
    # Nếu anh có base template cố định thì sửa lại chỗ này
    raise RuntimeError("Không tìm thấy file template cho Weekly Report trong thư mục Reports.")


# ====== EXCEL LAYOUT HELPERS ======

def resolve_summary_sheet(wb: Workbook) -> Worksheet:
    if SUMMARY_SHEET_NAME and SUMMARY_SHEET_NAME in wb.sheetnames:
        return wb[SUMMARY_SHEET_NAME]
    return wb.active


def read_block_weeks(wb: Workbook, default_year: Optional[int] = None) -> List[IsoWeek]:
    ws = resolve_summary_sheet(wb)
    weeks: list[IsoWeek] = []

    for i in range(TARGET_WEEKS_COUNT):
        row = TARGET_FIRST_WEEK_ROW + i
        value = ws.cell(row=row, column=TARGET_WEEK_COL).value
        if value is None or value == "":
            continue

        text = str(value).strip()

        year: Optional[int]
        week: int

        if "/" in text:
            w, y = text.split("/", 1)
            week = int(w.strip())
            year = int(y.strip())
        else:
            week = int(text.strip())
            year = default_year if default_year is not None else date.today().year

        weeks.append(IsoWeek(year=year, week=week))

    return weeks


def find_total_row(ws: Worksheet) -> Optional[int]:
    for r in range(DETAIL_FIRST_ROW, DETAIL_FIRST_ROW + DETAIL_MAX_ROWS):
        value = ws.cell(row=r, column=2).value
        if isinstance(value, str) and value.strip().upper() == "TOTAL":
            return r
    return None


def clear_target_rows(ws: Worksheet) -> None:
    for i in range(TARGET_WEEKS_COUNT):
        row = TARGET_FIRST_WEEK_ROW + i
        for col in range(3, TARGET_VOLUME_COL + 1):
            ws.cell(row=row, column=col, value=None)


def clear_detail_table(ws: Worksheet) -> None:
    total_row = find_total_row(ws)
    for r in range(DETAIL_FIRST_ROW, DETAIL_FIRST_ROW + DETAIL_MAX_ROWS):
        if total_row is not None and r == total_row:
            continue
        for c in range(2, 8):
            ws.cell(row=r, column=c, value=None)


def setup_new_block_layout(wb: Workbook, start_week: IsoWeek, block_size: int = TARGET_WEEKS_COUNT) -> None:
    ws = resolve_summary_sheet(wb)

    weeks: list[IsoWeek] = []
    current = start_week
    for i in range(block_size):
        if i == 0:
            weeks.append(current)
        else:
            weeks.append(weeks[-1].add_weeks(1))

    for i, wk in enumerate(weeks):
        row = TARGET_FIRST_WEEK_ROW + i
        ws.cell(row=row, column=TARGET_WEEK_COL, value=wk.week)

    clear_target_rows(ws)
    clear_detail_table(ws)


def is_week_empty(wb: Workbook, week: IsoWeek) -> bool:
    ws = resolve_summary_sheet(wb)
    block_weeks = read_block_weeks(wb, default_year=week.year)
    try:
        index = next(i for i, w in enumerate(block_weeks) if w.week == week.week and w.year == week.year)
    except StopIteration:
        raise ValueError(f"Tuần {week.label()} không nằm trong block 4 tuần của file báo cáo.")

    row = TARGET_FIRST_WEEK_ROW + index
    profit_cell = ws.cell(row=row, column=TARGET_PROFIT_COL).value
    volume_cell = ws.cell(row=row, column=TARGET_VOLUME_COL).value
    return (profit_cell in (None, "", 0, 0.0)) and (volume_cell in (None, "", 0, 0.0))


# ====== FILL DATA VÀO REPORT ======

def fill_week_data(wb: Workbook, week: IsoWeek, shipments: List[dict]) -> None:
    ws = resolve_summary_sheet(wb)

    total_row = find_total_row(ws)
    shipments_count = len(shipments)

    if total_row is None:
        # Không tìm thấy TOTAL, xoá sạch vùng và đặt TOTAL theo số dòng data
        clear_detail_table(ws)
        total_row = DETAIL_FIRST_ROW + shipments_count
    else:
        # Xoá data cũ nhưng giữ dòng TOTAL
        for r in range(DETAIL_FIRST_ROW, total_row):
            for c in range(2, 8):
                ws.cell(row=r, column=c, value=None)

        current_capacity = total_row - DETAIL_FIRST_ROW
        if shipments_count > current_capacity:
            extra = shipments_count - current_capacity
            ws.insert_rows(total_row, amount=extra)
            total_row += extra

    total_profit = 0.0
    total_volume = 0.0

    row = DETAIL_FIRST_ROW
    for i, item in enumerate(shipments, start=1):
        ws.cell(row=row, column=2, value=i)
        ws.cell(row=row, column=3, value=item["customer"])
        ws.cell(row=row, column=4, value=item["routing"])
        ws.cell(row=row, column=5, value=item["volume"])
        ws.cell(row=row, column=6, value=item["profit"])

        total_volume += float(item["volume"])
        total_profit += float(item["profit"])
        row += 1

    ws.cell(row=total_row, column=2, value="TOTAL")

    if shipments_count > 0:
        ws.cell(row=total_row, column=5, value=f"=SUM(E{DETAIL_FIRST_ROW}:E{total_row - 1})")
        ws.cell(row=total_row, column=6, value=f"=SUM(F{DETAIL_FIRST_ROW}:F{total_row - 1})")
    else:
        ws.cell(row=total_row, column=5, value=None)
        ws.cell(row=total_row, column=6, value=None)

    block_weeks = read_block_weeks(wb, default_year=week.year)
    try:
        index = next(i for i, w in enumerate(block_weeks) if w.week == week.week and w.year == week.year)
    except StopIteration:
        raise ValueError(f"Tuần {week.label()} không nằm trong block 4 tuần của file báo cáo.")

    summary_row = TARGET_FIRST_WEEK_ROW + index
    ws.cell(row=summary_row, column=TARGET_PROFIT_COL, value=total_profit)
    ws.cell(row=summary_row, column=TARGET_VOLUME_COL, value=total_volume)

    monday, sunday = iso_week_range(week.year, week.week)
    ws["E1"] = f"TUẦN: {week.week}"
    ws["B2"] = f"(Từ {monday.strftime('%d/%m')} đến {sunday.strftime('%d/%m')})"


def find_existing_file_for_week(week: IsoWeek) -> Optional[Path]:
    for path in reversed(iter_report_files()):
        wb = load_workbook(path, data_only=True)
        try:
            block_weeks = read_block_weeks(wb, default_year=week.year)
        finally:
            wb.close()

        for w in block_weeks:
            if w.week == week.week and w.year == week.year:
                return path

    return None


# ====== PUBLIC API: HÀM GỌI TỪ STREAMLIT ======

def generate_weekly_report(iso_year: int, iso_week: int, user_name: str) -> Path:
    week = IsoWeek(iso_year, iso_week)
    shipments = load_shipments_for_week(iso_year, iso_week)

    existing_path = find_existing_file_for_week(week)

    if existing_path is not None:
        wb = load_workbook(existing_path)
        try:
            if not is_week_empty(wb, week):
                raise ValueError(f"Report cho tuần {week.label()} đã có dữ liệu, không thể ghi đè.")
            fill_week_data(wb, week, shipments)
            wb.save(existing_path)
        finally:
            wb.close()

        new_name = build_report_filename(week, user_name)
        new_path = existing_path.with_name(new_name)
        if new_path != existing_path:
            if new_path.exists():
                new_path.unlink()
            existing_path.rename(new_path)
        return new_path

    template_source = choose_template_source()
    new_path = REPORTS_DIR / build_report_filename(week, user_name)
    new_path.write_bytes(template_source.read_bytes())

    wb = load_workbook(new_path)
    try:
        setup_new_block_layout(wb, week, TARGET_WEEKS_COUNT)
        fill_week_data(wb, week, shipments)
        wb.save(new_path)
    finally:
        wb.close()

    return new_path
