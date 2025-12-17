import os
import re
from datetime import datetime, date

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# === Config ===
base_dir = r"C:\Users\Nelson\OneDrive\Desktop\2. Areas\PricingSystem"

raw_dir = os.path.join(base_dir, "Raw")
data_dir = os.path.join(base_dir, "Data")

master_file = os.path.join(data_dir, "Master_FullPricing.xlsx")
puc_file = os.path.join(data_dir, "PUC_SOC.xlsx")
schedule_file = os.path.join(data_dir, "Schedule.xlsx")

os.makedirs(raw_dir, exist_ok=True)
os.makedirs(data_dir, exist_ok=True)

SOC_CARRIERS = ["CMA", "ONE", "YML"]


# ========= Utilities =========
def read_excel_safe(filepath, sheet_name="RATE", **kwargs):
    """
    Read Excel, prefer sheet 'RATE'. If not found, fall back to first sheet.
    """
    try:
        return pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl", **kwargs)
    except ValueError:
        return pd.read_excel(filepath, sheet_name=0, engine="openpyxl", **kwargs)


def get_col_safe(df, idx):
    """Get column by index, return None if out of range."""
    if idx is None:
        return None
    if idx < 0 or idx >= df.shape[1]:
        return None
    return df.iloc[:, idx]


def clean_amount_series(s):
    """Normalize amount string to numeric (remove commas)."""
    if s is None:
        return None
    s = s.astype(str).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(s, errors="coerce")


def format_short_date(series):
    """
    Format date to 'DD-MMM' (upper). If invalid, result is NaN.
    """
    dt = pd.to_datetime(series, errors="coerce")
    return dt.dt.strftime("%d-%b").str.upper()


def extract_version_from_filename(fname):
    """
    Extract version from file name.
    Example: 'FAK_US CANADA_ 2025 10 DEC NO 2.xlsx' -> '10DEC.NO2'
    """
    name = fname.upper().replace(" ", "").replace("_", "")
    match = re.search(r"(\d{1,2}[A-Z]{3})NO\.?\d+", name)
    if match:
        return match.group(0).replace(".", ".")
    return datetime.today().strftime("%d%b").upper() + ".NOX"


def write_version_sheet(writer, version_name, df_old_rate, raw_files):
    """
    Create a version sheet with version_name in the Excel workbook.
    """
    from openpyxl.worksheet.worksheet import Worksheet

    if version_name in writer.book.sheetnames:
        version_name += "_1"

    ws = writer.book.create_sheet(title=version_name)
    assert isinstance(ws, Worksheet) or True

    ws["A1"] = "Phiên bản bảng giá"
    ws["B1"] = version_name
    ws["A3"] = "Tổng số dòng"
    ws["B3"] = len(df_old_rate)
    ws["A4"] = "Ngày normalize"
    ws["B4"] = datetime.today().strftime("%d-%b-%Y")

    for idx, raw in enumerate(raw_files, start=6):
        ws[f"A{idx}"] = "RAW: " + str(raw)


def normalize_container(cont_series):
    cont = (
        cont_series.astype(str)
        .str.strip()
        .str.upper()
        .str.replace(" ", "", regex=False)
        .str.replace("'", "", regex=False)
    )
    return np.select(
        [
            cont.isin(["20", "20FT", "20DC", "20DV", "20GP"]),
            cont.isin(["40", "40FT", "40DC", "40DV", "40GP"]),
            cont.isin(["40HC", "40HQ", "40HCFT", "40HQFT"]),
        ],
        ["20GP", "40GP", "40HQ"],
        default=cont,
    )


# ========= Parser for FAK & ONE_SPECIAL RATE =========
def parse_fak_or_fix(raw_df, rate_type, source_file):
    if raw_df.shape[0] < 3:
        return pd.DataFrame()

    df = raw_df.iloc[2:].reset_index(drop=True)

    POL = get_col_safe(df, 0)
    POD = get_col_safe(df, 1)

    Place = get_col_safe(df, 2)
    if Place is not None:
        Place = Place.astype(str).str.strip()

    RoutingNote = get_col_safe(df, 3)
    Carrier = get_col_safe(df, 5)
    Eff = get_col_safe(df, 6)
    Exp = get_col_safe(df, 7)
    Commodity = get_col_safe(df, 9) if rate_type == "FAK" else None
    Contract = get_col_safe(df, 11)

    c20 = get_col_safe(df, 12)
    c40 = get_col_safe(df, 13)
    c40hq = get_col_safe(df, 14)
    c45 = get_col_safe(df, 15)
    c40nor = get_col_safe(df, 16)

    puc20_raw = get_col_safe(df, 38)
    puc40_raw = get_col_safe(df, 39)
    puc40hq_raw = get_col_safe(df, 40)
    puc45_raw = get_col_safe(df, 41)

    c20 = clean_amount_series(c20)
    c40 = clean_amount_series(c40)
    c40hq = clean_amount_series(c40hq)
    c45 = clean_amount_series(c45)
    c40nor = clean_amount_series(c40nor)

    puc20 = clean_amount_series(puc20_raw)
    puc40 = clean_amount_series(puc40_raw)
    puc40hq = clean_amount_series(puc40hq_raw)
    puc45 = clean_amount_series(puc45_raw)

    price_df = pd.DataFrame(
        {
            "POL": POL,
            "POD": POD,
            "PlaceOfDelivery": Place,
            "RoutingNote": RoutingNote,
            "Carrier": Carrier,
            "EffectiveDate": Eff,
            "ExpirationDate": Exp,
            "ContractIdentifier": Contract,
            "CommodityType": Commodity,
            "20GP": c20,
            "40GP": c40,
            "40HQ": c40hq,
            "45HQ": c45,
            "40NOR": c40nor,
        }
    )

    price_df["PUC20"] = puc20 if puc20 is not None else 0
    price_df["PUC40"] = puc40 if puc40 is not None else 0
    price_df["PUC40HQ"] = puc40hq if puc40hq is not None else 0
    price_df["PUC45"] = puc45 if puc45 is not None else 0

    mask_carrier = price_df["Carrier"].astype(str).str.upper().isin(
        [c.upper() for c in SOC_CARRIERS]
    )
    mask_soc_note = price_df["RoutingNote"].astype(str).str.upper().str.contains(
        "SOC", na=False
    )
    mask_apply_puc = mask_carrier & mask_soc_note

    if mask_apply_puc.any():
        for cont_col, puc_col in [
            ("20GP", "PUC20"),
            ("40GP", "PUC40"),
            ("40HQ", "PUC40HQ"),
        ]:
            if cont_col in price_df.columns:
                price_df.loc[mask_apply_puc, cont_col] = (
                    price_df.loc[mask_apply_puc, cont_col].fillna(0)
                    - price_df.loc[mask_apply_puc, puc_col].fillna(0)
                )

    price_df = price_df.drop(columns=["PUC20", "PUC40", "PUC40HQ", "PUC45"], errors="ignore")

    value_cols = ["20GP", "40GP", "40HQ", "45HQ", "40NOR"]
    value_cols = [c for c in value_cols if c in price_df.columns]

    df_out = (
        price_df.melt(
            id_vars=[
                "POL",
                "POD",
                "PlaceOfDelivery",
                "RoutingNote",
                "Carrier",
                "EffectiveDate",
                "ExpirationDate",
                "ContractIdentifier",
                "CommodityType",
            ],
            value_vars=value_cols,
            var_name="ContainerType",
            value_name="Amount",
        )
        .dropna(subset=["Amount"])
    )

    if rate_type == "ONE_SPECIAL RATE":
        df_out["CommodityType"] = "FIX RATE"

    mask_one = df_out["Carrier"].astype(str).str.upper().str.contains("ONE", na=False)
    if mask_one.any():
        comm = df_out.loc[mask_one, "CommodityType"].astype(str)

        m_gar = comm.str.contains("GARMENT", case=False, na=False)
        comm.loc[m_gar] = "FAK: TPE1 - FAK Straight"

        m1 = comm.str.contains("FAK: TPE1 - FAK STRAIGHT", case=False, na=False)
        comm.loc[m1] = "FAK: TPE1 - FAK Straight"

        m2 = comm.str.contains("REEFER FAK", case=False, na=False)
        comm.loc[m2] = "REEFER FAK"

        m3 = comm.str.contains("SHORT TERM GDSM", case=False, na=False)
        comm.loc[m3] = "SHORT TERM GDSM"

        m4 = comm.str.contains("TPE9", case=False, na=False) & comm.str.contains(
            "GROUP SOC", case=False, na=False
        )
        comm.loc[m4] = "S1– TPE9 – Group SOC"

        df_out.loc[mask_one, "CommodityType"] = comm

    df_out["Amount"] = clean_amount_series(df_out["Amount"])
    df_out["RateType"] = rate_type
    df_out["SourceFile"] = source_file
    df_out = df_out.drop_duplicates()

    final_cols = [
        "POL",
        "POD",
        "PlaceOfDelivery",
        "RoutingNote",
        "Carrier",
        "EffectiveDate",
        "ExpirationDate",
        "ContractIdentifier",
        "ContainerType",
        "Amount",
        "CommodityType",
        "RateType",
        "SourceFile",
    ]
    return df_out[final_cols]


# ========= Parser for HPL_SCFI =========
def parse_scfi(raw_df, source_file):
    if raw_df.shape[0] < 3:
        return pd.DataFrame()

    df = raw_df.iloc[2:].reset_index(drop=True)

    POL = get_col_safe(df, 0)
    POD = get_col_safe(df, 1)

    Place = get_col_safe(df, 2)
    if Place is not None:
        Place = Place.astype(str).str.strip()

    Eff = get_col_safe(df, 3)
    Exp = get_col_safe(df, 4)
    c20 = get_col_safe(df, 5)
    c40 = get_col_safe(df, 6)
    c40hq = get_col_safe(df, 7)

    price_df = pd.DataFrame(
        {
            "POL": POL,
            "POD": POD,
            "PlaceOfDelivery": Place,
            "RoutingNote": None,
            "Carrier": "HPL",
            "EffectiveDate": Eff,
            "ExpirationDate": Exp,
            "ContractIdentifier": None,
            "CommodityType": None,
            "20GP": c20,
            "40GP": c40,
            "40HQ": c40hq,
        }
    )

    df_out = (
        price_df.melt(
            id_vars=[
                "POL",
                "POD",
                "PlaceOfDelivery",
                "RoutingNote",
                "Carrier",
                "EffectiveDate",
                "ExpirationDate",
                "ContractIdentifier",
                "CommodityType",
            ],
            value_vars=["20GP", "40GP", "40HQ"],
            var_name="ContainerType",
            value_name="Amount",
        )
        .dropna(subset=["Amount"])
    )

    df_out["Amount"] = clean_amount_series(df_out["Amount"])
    df_out["RateType"] = "HPL_SCFI"
    df_out["SourceFile"] = source_file

    final_cols = [
        "POL",
        "POD",
        "PlaceOfDelivery",
        "RoutingNote",
        "Carrier",
        "EffectiveDate",
        "ExpirationDate",
        "ContractIdentifier",
        "ContainerType",
        "Amount",
        "CommodityType",
        "RateType",
        "SourceFile",
    ]
    return df_out[final_cols]


# ========= Detect rate type from file name =========
def detect_rate_type_from_name(fname):
    name_upper = fname.upper()

    if "SCFI" in name_upper:
        return "HPL_SCFI"
    if "FAK" in name_upper:
        return "FAK"

    if "ONE_SPECIAL RATE" in name_upper:
        return "ONE_SPECIAL RATE"
    if "ONE_SPECIAL" in name_upper and "RATE" in name_upper:
        return "ONE_SPECIAL RATE"

    if "ONE_FIX" in name_upper:
        return "ONE_SPECIAL RATE"
    if "ONE FIX" in name_upper:
        return "ONE_SPECIAL RATE"
    if "ONE-FIX" in name_upper:
        return "ONE_SPECIAL RATE"
    if "FIX" in name_upper and "ONE" in name_upper:
        return "ONE_SPECIAL RATE"
    if "FIX" in name_upper:
        return "ONE_SPECIAL RATE"

    return None


# ========= Load PUC =========
def load_puc():
    if not os.path.exists(puc_file):
        print("[!] Missing PUC_SOC.xlsx at: {0} -> skip PUC.".format(puc_file))
        return None

    df = pd.read_excel(puc_file, sheet_name="PUC_SOC")
    df.columns = [str(c).strip() for c in df.columns]

    required = ["PlaceOfDelivery", "20DC", "40HC"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError("PUC_SOC.xlsx missing columns: {0}".format(missing))

    for col in ["20DC", "40HC"]:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
            .replace({"": pd.NA, "TBA": pd.NA, "N/A": pd.NA})
            .astype(float)
        )

    df["CityKey"] = df["PlaceOfDelivery"].astype(str).str.upper().str.strip()
    return df


def build_city_key_for_master(pod, puc_cities_upper):
    if pd.isna(pod):
        return ""
    up = str(pod).upper()
    matches = [city for city in puc_cities_upper if city in up]
    if matches:
        return sorted(matches, key=len, reverse=True)[0]
    base = up.split("(")[0]
    base = base.split(",")[0].strip()
    return base


def apply_puc_to_df(df_master):
    puc_df = load_puc()
    if puc_df is None:
        return df_master

    df = df_master.copy()
    df["ContainerNorm"] = normalize_container(df["ContainerType"])

    if "PlaceOfDelivery" not in df.columns:
        raise ValueError("Master missing PlaceOfDelivery column")

    puc_cities_upper = puc_df["CityKey"].dropna().unique().tolist()
    df["CityKey"] = df["PlaceOfDelivery"].apply(
        lambda x: build_city_key_for_master(x, puc_cities_upper)
    )

    df = df.merge(
        puc_df[["CityKey", "20DC", "40HC"]],
        on="CityKey",
        how="left",
        suffixes=("", "_PUC"),
    )

    df["PUC_selected"] = np.where(
        df["ContainerNorm"] == "20GP",
        df["20DC"],
        np.where(df["ContainerNorm"].isin(["40GP", "40HQ"]), df["40HC"], np.nan),
    )

    mask_carrier = df["Carrier"].astype(str).str.upper().isin(
        [c.upper() for c in SOC_CARRIERS]
    )
    mask_soc_note = df["RoutingNote"].astype(str).str.upper().str.contains(
        "SOC", na=False
    )
    mask_apply = mask_carrier & mask_soc_note

    puc_series = df.loc[mask_apply, "PUC_selected"].fillna(0)
    df.loc[mask_apply, "Amount"] = df.loc[mask_apply, "Amount"] + puc_series

    drop_cols = ["ContainerNorm", "CityKey", "20DC", "40HC", "PUC_selected"]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns])

    return df


# ========= Normalize POD & PlaceOfDelivery =========
def load_port_mapping():
    port_file = os.path.join(data_dir, "Port_Code_Mapping_Final.xlsx")
    if not os.path.exists(port_file):
        print("[!] Missing Port_Code_Mapping_Final.xlsx at: {0}".format(port_file))
        return {}

    df = pd.read_excel(port_file)
    df.columns = [str(c).strip().upper() for c in df.columns]
    df = df.dropna(subset=["PORTNAME", "PORTCODE"])
    df["PORTNAME"] = df["PORTNAME"].astype(str).str.upper().str.strip()
    df["PORTCODE"] = df["PORTCODE"].astype(str).str.strip()
    return dict(zip(df["PORTNAME"], df["PORTCODE"]))


def normalize_pod_column(df, pod_map):
    df = df.copy()
    df["POD"] = df["POD"].apply(
        lambda x: pod_map.get(str(x).strip().upper(), str(x).strip())
        if pd.notna(x)
        else x
    )
    return df


def normalize_place_of_delivery_column(df):
    df = df.copy()
    df["PlaceOfDelivery"] = df["PlaceOfDelivery"].astype(str).str.upper().str.strip()
    return df


# ========= Normalize commodity by carrier =========
def normalize_commodity(df_master):
    if "CommodityType" not in df_master.columns or "Carrier" not in df_master.columns:
        return df_master

    df = df_master.copy()

    carrier_upper = df["Carrier"].astype(str).str.upper()
    comm = df["CommodityType"].astype(str)

    mask_cosco = carrier_upper.str.contains("COSCO", na=False)

    mask_cosco_fak_ex = mask_cosco & comm.str.contains(
        "FAK (Excluding Garment)", case=False, na=False, regex=False
    )
    df.loc[mask_cosco_fak_ex, "CommodityType"] = "FAK"

    mask_cosco_gar = mask_cosco & comm.str.contains(
        "Garments/Textile/Consol", case=False, na=False, regex=False
    )
    df.loc[mask_cosco_gar, "CommodityType"] = "GARMENT"

    mask_emc = carrier_upper.str.contains("EMC", na=False)
    mask_emc_rate1 = mask_emc & comm.str.contains(
        "RATE 1 - GENERAL CARGO", case=False, na=False, regex=False
    )
    df.loc[mask_emc_rate1, "CommodityType"] = "RATE 1"

    mask_hpl = carrier_upper.str.contains("HPL", na=False)

    mask_hpl_fak_inc = mask_hpl & comm.str.contains(
        "FAK INCLUDING GARMENT", case=False, na=False, regex=False
    )
    df.loc[mask_hpl_fak_inc, "CommodityType"] = "FAK"

    mask_yml = carrier_upper.str.contains("YML", na=False)
    pattern_yml_fak = "FAK (NON-HAZ, EXCLUDING REEFER/ SHIPS/ BOATS/ VEHICLES/ CARS)"

    mask_yml_groupA = (
        mask_yml
        & comm.str.contains("GROUP A", case=False, na=False, regex=False)
        & comm.str.contains(pattern_yml_fak, case=False, na=False, regex=False)
    )
    df.loc[mask_yml_groupA, "CommodityType"] = "GROUP A"

    mask_yml_fak = (
        mask_yml
        & comm.str.contains(pattern_yml_fak, case=False, na=False, regex=False)
        & ~mask_yml_groupA
    )
    df.loc[mask_yml_fak, "CommodityType"] = "FAK"

    return df


# ========= Long -> Wide =========
def make_horizontal_output(df_long):
    if df_long.empty:
        return df_long.copy()

    df_long = df_long.copy()

    wanted_keys = [
        "POL",
        "POD",
        "PlaceOfDelivery",
        "RoutingNote",
        "Carrier",
        "EffectiveDate",
        "ExpirationDate",
        "ContractIdentifier",
        "CommodityType",
        "RateType",
    ]
    index_cols = [c for c in wanted_keys if c in df_long.columns]

    extra_keys = [
        c
        for c in df_long.columns
        if c not in index_cols and c not in ["ContainerType", "Amount"]
    ]
    index_cols = index_cols + extra_keys

    for col in index_cols:
        df_long[col] = df_long[col].fillna("")

    wide = (
        df_long.pivot_table(
            index=index_cols,
            columns="ContainerType",
            values="Amount",
            aggfunc="first",
        )
        .reset_index()
    )

    if wide.columns.name:
        wide.columns.name = None

    cont_order = ["20GP", "40GP", "40HQ", "45HQ", "40NOR"]
    key_cols = [c for c in wide.columns if c not in cont_order]
    other = [c for c in wide.columns if c not in key_cols and c not in cont_order]

    final_cols = key_cols + [c for c in cont_order if c in wide.columns] + other
    return wide[final_cols]


# ========= Snapshot current / previous =========
def split_current_and_previous_long(
            df_long: pd.DataFrame,
            include_expired: bool = False,
            cutoff_date: date | None = None,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Từ FULL history (df_long), tính ra:
          - df_current_long: các dòng dùng cho sheet Master
                + filter theo ExpirationDate >= cutoff_date (hoặc include_expired=True thì không filter)
          - df_prev_long: 1 dòng "previous" (giá cũ) align với từng dòng trong df_current_long
                + dùng để tính DELTA_* trên Master
        Lưu ý:
          - Hàm này KHÔNG drop history; toàn bộ lịch sử vẫn còn trong df_long
            (sẽ được pivot thành sheet Old_Rate ở combine_all()).
          - Previous = bản ghi liền trước theo thời gian trong cùng group lane+carrier+RateType+ContainerType.
        """
        if df_long.empty:
            return df_long.copy(), df_long.copy()

        df = df_long.copy()

        group_cols = [
            "POL",
            "POD",
            "PlaceOfDelivery",
            "RoutingNote",
            "Carrier",
            "ContractIdentifier",
            "CommodityType",
            "RateType",
            "ContainerType",
        ]
        group_cols = [c for c in group_cols if c in df.columns]

        # Chuẩn hóa ngày để sort + filter
        df["_EffDateDT"] = pd.to_datetime(df.get("EffectiveDate"), errors="coerce")
        df["_ExpDateDT"] = pd.to_datetime(df.get("ExpirationDate"), errors="coerce").dt.date

        df = df.sort_values(group_cols + ["_EffDateDT", "_ExpDateDT"])

        df["_order"] = df.groupby(group_cols).cumcount()
        df["_max_order"] = df.groupby(group_cols)["_order"].transform("max")

        # Xác định cutoff cho Master
        if cutoff_date is None:
            cutoff_date = date.today()

        if include_expired:
            # Master show toàn bộ (không filter theo ExpDate)
            mask_valid_for_master = pd.Series(True, index=df.index)
        else:
            mask_valid_for_master = df["_ExpDateDT"].isna() | (df["_ExpDateDT"] >= cutoff_date)

        df_current_long = df[mask_valid_for_master].copy()

        # Tính previous cho từng dòng current bằng cách dịch _order
        # Ý tưởng: dòng previous có cùng group_cols và _order = current._order - 1
        prev_map = df[group_cols + ["_order", "Amount"]].copy()
        prev_map["_order"] = prev_map["_order"] + 1  # dịch lên 1 để merge được "order-1"
        prev_map = prev_map.rename(columns={"Amount": "Amount_prev"})

        df_current_long = df_current_long.merge(
            prev_map,
            on=group_cols + ["_order"],
            how="left",
        )

        # Tạo df_prev_long: cùng key với current nhưng Amount = Amount_prev
        df_prev_long = df_current_long.copy()
        if "Amount_prev" in df_prev_long.columns:
            df_prev_long["Amount"] = df_prev_long["Amount_prev"]
        # Bỏ những dòng không có previous (Amount_prev NaN) để tránh noise
        df_prev_long = df_prev_long.dropna(subset=["Amount"])

        # Dọn helper columns
        drop_cols = ["_EffDateDT", "_ExpDateDT", "_order", "_max_order", "Amount_prev"]
        df_current_long = df_current_long.drop(columns=[c for c in drop_cols if c in df_current_long.columns],
                                               errors="ignore")
        df_prev_long = df_prev_long.drop(columns=[c for c in drop_cols if c in df_prev_long.columns], errors="ignore")

        return df_current_long, df_prev_long


def add_delta_display_columns(master_current, master_previous):
    if master_current.empty:
        return master_current.copy()

    df = master_current.copy()
    container_cols = ["20GP", "40GP", "40HQ", "45HQ", "40NOR"]

    base_key_cols = [c for c in df.columns if c not in container_cols]
    join_key_cols = [c for c in base_key_cols if c not in ["EffectiveDate", "ExpirationDate"]]

    prev_available = [c for c in container_cols if c in master_previous.columns]
    prev_rename = {c: c + "_OLD" for c in prev_available}

    if (not master_previous.empty) and prev_available:
        master_prev_trim = master_previous[join_key_cols + prev_available]
        master_prev_trim = master_prev_trim.rename(columns=prev_rename)
        df = df.merge(master_prev_trim, on=join_key_cols, how="left")
    else:
        for c in prev_available:
            df[c + "_OLD"] = np.nan

    def format_delta_icon(delta):
        if pd.isna(delta):
            return ""
        try:
            delta_val = float(delta)
        except Exception:
            return ""

        if np.isclose(delta_val, 0):
            icon = "↔️"
            mag = 0.0
        elif delta_val > 0:
            icon = "⬆️"
            mag = delta_val
        else:
            icon = "⬇️"
            mag = abs(delta_val)

        if float(mag).is_integer():
            mag_str = str(int(mag))
        else:
            mag_str = str(round(mag, 2))

        return f"{icon} {mag_str}"

    view_cols = []
    old_cols_to_drop = []
    numeric_delta_cols = []

    for cont in container_cols:
        price_col = cont
        old_col = cont + "_OLD"
        if price_col not in df.columns or old_col not in df.columns:
            continue

        delta_col = "DELTA_" + cont
        view_col = cont + "_VIEW"

        df[delta_col] = df[price_col] - df[old_col]
        df[view_col] = df[delta_col].apply(format_delta_icon)

        view_cols.append(view_col)
        old_cols_to_drop.append(old_col)
        numeric_delta_cols.append(delta_col)

    # ---- PHẦN QUAN TRỌNG: giữ DELTA_* , chỉ drop *_OLD ----
    base_cols = [c for c in master_current.columns if c not in container_cols]
    price_cols = [c for c in container_cols if c in master_current.columns]

    extra_cols = [
        c
        for c in df.columns
        if c
        not in base_cols + price_cols + old_cols_to_drop + numeric_delta_cols + view_cols
    ]

    # chỉ xoá cột *_OLD dùng tạm cho join
    drop_cols = old_cols_to_drop
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")

    # Master sẽ có: key + giá + VIEW + DELTA + extra
    ordered_cols = base_cols + price_cols + view_cols + numeric_delta_cols + extra_cols
    return df[ordered_cols]

def pixels_to_width(pixels):
    return max((pixels - 5) / 7.0, 0.0)


def set_master_column_widths(ws):
    pixel_map = {
        "A": 80,
        "B": 141,
        "C": 141,
        "D": 125,
        "E": 85,
        "F": 112,
        "G": 112,
        "H": 146,
        "I": 150,
        "J": 80,
        "K": 80,
        "L": 80,
        "M": 80,
        "N": 80,
        "O": 80,
    }

    for col, px in pixel_map.items():
        ws.column_dimensions[col].width = pixels_to_width(px)


def apply_delta_icon_format(ws, view_col_letter, start_row, end_row):
    if not view_col_letter:
        return
    if end_row < start_row:
        return

    cell_range = "{0}{1}:{0}{2}".format(view_col_letter, start_row, end_row)

    dark_green_font = Font(color="FF008000")   # xanh đậm
    light_green_font = Font(color="FF00B050")  # xanh nhạt
    orange_font = Font(color="FFFFC000")       # cam
    red_font = Font(color="FFFF0000")          # đỏ
    neutral_font = Font(color="FF808080")      # xám

    col = view_col_letter

    # Giảm mạnh: ⬇️ và |delta| >= 200
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=AND(LEFT(${col}{start_row},1)="⬇", VALUE(MID(${col}{start_row},3,99))>=200)'],
            font=dark_green_font,
        ),
    )

    # Giảm nhẹ: ⬇️ và |delta| < 200
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=AND(LEFT(${col}{start_row},1)="⬇", VALUE(MID(${col}{start_row},3,99))<200)'],
            font=light_green_font,
        ),
    )

    # Tăng mạnh: ⬆️ và delta >= 200
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=AND(LEFT(${col}{start_row},1)="⬆", VALUE(MID(${col}{start_row},3,99))>=200)'],
            font=red_font,
        ),
    )

    # Tăng nhẹ: ⬆️ và delta < 200
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=AND(LEFT(${col}{start_row},1)="⬆", VALUE(MID(${col}{start_row},3,99))<200)'],
            font=orange_font,
        ),
    )

    # Extend: ↔️
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=LEFT(${col}{start_row},1)="↔"'],
            font=neutral_font,
        ),
    )


def normalize_file(filepath):
    fname = os.path.basename(filepath)
    print("\n[+] Processing RAW: {0}".format(fname))

    rate_type = detect_rate_type_from_name(fname)
    if rate_type is None:
        print("    -> Cannot detect RateType from file name, skip.")
        return pd.DataFrame()

    file_size = os.path.getsize(filepath) / (1024 * 1024)
    if file_size > 10:
        chunks = []
        excel_obj = pd.read_excel(
            filepath,
            sheet_name=None,
            header=None,
            engine="openpyxl",
        )
        for _, chunk in excel_obj.items():
            chunks.append(chunk)
        raw_df = pd.concat(chunks)
    else:
        raw_df = read_excel_safe(filepath, header=None)

    if raw_df.empty:
        print("    -> Empty file, skip.")
        return pd.DataFrame()

    if rate_type in ["FAK", "ONE_SPECIAL RATE"]:
        df_out = parse_fak_or_fix(raw_df, rate_type, fname)
    else:
        df_out = parse_scfi(raw_df, fname)

    if df_out.empty:
        print("    -> No valid rows parsed.")
        return pd.DataFrame()

    print("    -> Normalized (long): {0} rows.".format(len(df_out)))
    return df_out


def combine_all(list_df, include_expired: bool = False, cutoff_date: date | None = None):
    """
    Gộp tất cả DataFrame long, chuẩn hóa, rồi tạo:
      - Master (ngang): chỉ các dòng còn hiệu lực (ExpirationDate >= cutoff_date
        hoặc ExpirationDate null, trừ khi include_expired=True) + cột DELTA_*_DISPLAY
        thể hiện biến động so với kỳ liền trước.
      - Old_Rate (ngang): FULL history đã normalize (không dùng snapshot 2 kỳ),
        dùng làm nguồn để review / tra cứu giá cũ.
    """
    if not list_df:
        print("No valid data to combine.")
        return

    # 1) Gộp full history từ tất cả file RAW (long)
    master = pd.concat(list_df, ignore_index=True)

    print("\n[STATS] Rows by RateType (FULL history, before normalization):")
    if "RateType" in master.columns:
        print(master["RateType"].value_counts())

    # 2) Áp PUC + normalize POD / Place / Commodity trên FULL history
    master_full = apply_puc_to_df(master)

    master_full["PlaceOfDelivery"] = master_full["PlaceOfDelivery"].astype(str).str.strip()

    pod_map = load_port_mapping()
    master_full = normalize_pod_column(master_full, pod_map)
    master_full = normalize_place_of_delivery_column(master_full)
    master_full = normalize_commodity(master_full)

    # Giữ SourceFile cho debugging ngoài, nhưng không cần cho Master/Old_Rate
    master_full_no_src = master_full.drop(columns=["SourceFile"], errors="ignore")

    # 3) Old_Rate: FULL history (ngang) - KHÔNG snapshot, KHÔNG filter Expiration
    old_rate_wide = make_horizontal_output(master_full_no_src)

    print("\n[OLD_RATE] Full normalized history rows (long): {0}".format(len(master_full_no_src)))
    print("[OLD_RATE] Full normalized history rows (wide): {0}".format(len(old_rate_wide)))

    # 4) Tính current/previous cho Master dựa trên FULL history
    df_current_long, df_prev_long = split_current_and_previous_long(
        master_full_no_src,
        include_expired=include_expired,
        cutoff_date=cutoff_date,
    )

    print(
        "\n[MASTER] Rows for Master (current, after Expiration filter): {0}".format(
            len(df_current_long)
        )
    )
    print("[MASTER] Rows with previous available for DELTA: {0}".format(len(df_prev_long)))

    # Format ngày cho mục đích hiển thị (chỉ Master)
    for df_tmp in (df_current_long, df_prev_long):
        if df_tmp.empty:
            continue
        df_tmp["EffectiveDate"] = format_short_date(df_tmp["EffectiveDate"])
        df_tmp["ExpirationDate"] = format_short_date(df_tmp["ExpirationDate"])

    # 5) Pivot sang ngang cho Master (current + previous)
    master_current = make_horizontal_output(df_current_long)
    master_previous = make_horizontal_output(df_prev_long)

    # 6) Thêm cột DELTA_*_DISPLAY (icon + number) vào Master
    master_with_delta = add_delta_display_columns(master_current, master_previous)

    # 7) Ghi Excel
    with pd.ExcelWriter(master_file, engine="openpyxl") as writer:
        # Sheet Master: giá hiện tại + delta
        master_with_delta.to_excel(writer, index=False, sheet_name="Master")
        ws = writer.sheets["Master"]

        # Tô màu header
        header_fill = PatternFill(
            start_color="CCFFCC",
            end_color="CCFFCC",
            fill_type="solid",
        )
        for cell in ws[1]:
            cell.fill = header_fill

        # Set width các cột chính
        set_master_column_widths(ws)

        data_start_row = 2
        data_end_row = ws.max_row

        # Áp conditional formatting cho các cột DELTA_*_DISPLAY
        header_row = ws[1]
        header_to_col = {cell.value: get_column_letter(cell.column) for cell in header_row}

        container_cols = ["20GP", "40GP", "40HQ", "45HQ", "40NOR"]

        for cont in container_cols:
            view_header = f"{cont}_VIEW"
            view_col_letter = header_to_col.get(view_header)

            if view_col_letter:
                apply_delta_icon_format(
                    ws,
                    view_col_letter=view_col_letter,
                    start_row=data_start_row,
                    end_row=data_end_row,
                )

        # Ẩn các cột DELTA_* để sales chỉ thấy VIEW
        for cont in container_cols:
            delta_header = f"DELTA_{cont}"
            delta_col_letter = header_to_col.get(delta_header)
            if delta_col_letter:
                ws.column_dimensions[delta_col_letter].hidden = True

        # Sheet Old_Rate: FULL history (ngang)
        old_rate_wide.to_excel(writer, index=False, sheet_name="Old_Rate")

        # Version sheet dựa trên file FAK mới nhất trong Raw
        fak_files = [f for f in os.listdir(raw_dir) if "FAK" in f.upper()]
        if fak_files:
            latest_fak = sorted(fak_files)[-1]
            version_name = extract_version_from_filename(latest_fak)
            write_version_sheet(writer, version_name, master_full_no_src, fak_files)
            print("[VERSION] Added version sheet: {0}".format(version_name))
        else:
            print("[VERSION] No FAK file found in Raw/ to create version sheet.")

        # Sheet Schedule (giữ nguyên như trước)
        if os.path.exists(schedule_file):
            try:
                df_sched = pd.read_excel(schedule_file)
                df_sched.to_excel(writer, index=False, sheet_name="Schedule")
                print("[SCHEDULE] Attached 'Schedule' sheet from: {0}".format(schedule_file))
            except Exception as e:
                print("[SCHEDULE] Error when reading/writing Schedule.xlsx: {0}".format(e))
        else:
            print("[SCHEDULE] Missing Schedule.xlsx at: {0} -> skip.".format(schedule_file))

    print("\n[DONE] MASTER FILE: {0}".format(master_file))
    print("    -> Master rows (current view) : {0}".format(len(master_current)))
    print("    -> Old_Rate rows (full history): {0}".format(len(old_rate_wide)))

def main(include_expired: bool = False, cutoff_date: date | None = None):
    """
    Chạy normalize toàn bộ file trong thư mục Raw:
      - Đọc tất cả .xlsx trong raw_dir
      - Gọi normalize_file() cho từng file -> list_df (dạng long)
      - Gọi combine_all() để tạo Master / Old_Rate / version / Schedule
    """
    if not os.path.isdir(raw_dir):
        print(f"Raw folder does not exist: {raw_dir}")
        return

    files = [f for f in os.listdir(raw_dir) if f.lower().endswith(".xlsx")]
    if not files:
        print("No .xlsx file found in Raw folder.")
        return

    all_normalized: list[pd.DataFrame] = []
    for f in files:
        filepath = os.path.join(raw_dir, f)
        df_norm = normalize_file(filepath)
        if not df_norm.empty:
            all_normalized.append(df_norm)

    combine_all(all_normalized, include_expired=include_expired, cutoff_date=cutoff_date)


if __name__ == "__main__":
    main()