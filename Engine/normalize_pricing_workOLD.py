import os
import pandas as pd
from openpyxl.styles import PatternFill
import numpy as np
from datetime import datetime, timedelta, date

# === Cấu hình ===
base_dir = r"C:\Users\Nelson\OneDrive\Desktop\2. Areas\PricingSystem"

raw_dir = os.path.join(base_dir, "Raw")
data_dir = os.path.join(base_dir, "Data")

# MASTER FILE BÂY GIỜ LƯU TRONG THƯ MỤC Data
master_file = os.path.join(data_dir, "Master_FullPricing.xlsx")

# File PUC_SOC.xlsx nằm trong thư mục Data
puc_file = os.path.join(data_dir, "PUC_SOC.xlsx")

# ✅ File lịch tàu riêng (Schedule.xlsx) – sẽ được gắn vào sheet "Schedule"
schedule_file = os.path.join(data_dir, "Schedule.xlsx")

os.makedirs(raw_dir, exist_ok=True)
os.makedirs(data_dir, exist_ok=True)

# Các hãng áp dụng PUC cho giá SOC
SOC_CARRIERS = ["CMA", "ONE", "YML"]


# ========= HÀM TIỆN ÍCH =========
def read_excel_safe(filepath, sheet_name="RATE", **kwargs):
    """
    Đọc Excel, ưu tiên sheet 'RATE'. Nếu không có thì đọc sheet đầu tiên (index 0).
    """
    try:
        return pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl", **kwargs)
    except ValueError:
        return pd.read_excel(filepath, sheet_name=0, engine="openpyxl", **kwargs)


def get_col_safe(df, idx):
    """Lấy cột theo index, nếu không có thì trả None."""
    if idx is None:
        return None
    if idx < 0 or idx >= df.shape[1]:
        return None
    return df.iloc[:, idx]


def clean_amount_series(s):
    """Chuẩn hóa chuỗi Amount về số (bỏ dấu phẩy)."""
    if s is None:
        return None
    s = (
        s.astype(str)
        .str.replace(",", "", regex=False)  # bỏ thousand separator
        .str.strip()
    )
    return pd.to_numeric(s, errors="coerce")


def format_short_date(series):
    """
    Chuẩn hóa ngày về dạng '14-MAR' (DD-MMM, viết hoa, không năm, không giờ).
    Nếu không đọc được ngày → NaT -> trả về NaN.
    """
    dt = pd.to_datetime(series, errors="coerce")
    return dt.dt.strftime("%d-%b").str.upper()
import re

def extract_version_from_filename(fname):
    """
    Trích version từ tên file dạng 'FAK_US CANADA_ 2025 10 DEC NO 2.xlsx'
    -> Trả về '10DEC.NO2'
    """
    name = fname.upper().replace(" ", "").replace("_", "")
    match = re.search(r"(\d{1,2}[A-Z]{3})NO\.?\d+", name)
    if match:
        return match.group(0).replace(".", ".")
    return datetime.today().strftime("%d%b").upper() + ".NOX"

def write_version_sheet(writer, version_name, df_master_long, raw_files: list[str]):
    """
    Ghi thêm 1 sheet đặc biệt chứa tên version (vd '10DEC.NO2') vào Excel.
    """
    from openpyxl.worksheet.worksheet import Worksheet

    if version_name in writer.book.sheetnames:
        version_name += "_1"  # tránh trùng

    ws: Worksheet = writer.book.create_sheet(title=version_name)
    ws["A1"] = "Phiên bản bảng giá"
    ws["B1"] = version_name
    ws["A3"] = "Tổng số dòng"
    ws["B3"] = len(df_master_long)
    ws["A4"] = "Ngày normalize"
    ws["B4"] = datetime.today().strftime("%d-%b-%Y")

    for idx, raw in enumerate(raw_files, start=6):
        ws[f"A{idx}"] = f"RAW: {raw}"


def normalize_container(cont_series: pd.Series) -> pd.Series:
    """
    Vectorized version: Xử lý toàn series thay vì per-value.
    """
    cont = (
        cont_series.astype(str)
        .str.strip()
        .str.upper()
        .str.replace(" ", "", regex=False)
        .str.replace("'", "", regex=False)
    )
    return np.select(
        [
            cont.isin(["20", "20FT", "20DC", "20DV", "20GP"]),
            cont.isin(["40", "40FT", "40DC", "40DV", "40GP"]),
            cont.isin(["40HC", "40HQ", "40HCFT", "40HQFT"]),
        ],
        ["20GP", "40GP", "40HQ"],
        default=cont,
    )


# ========= PARSER CHO FAK & ONE_SPECIAL RATE =========
def parse_fak_or_fix(raw_df, rate_type, source_file):
    """
    Cấu trúc FAK & ONE_SPECIAL RATE theo mapping:
    ...
    (giữ nguyên mô tả cũ)
    """
    if raw_df.shape[0] < 3:
        return pd.DataFrame()

    # Bỏ 2 hàng header, dữ liệu bắt đầu từ dòng 3
    df = raw_df.iloc[2:].reset_index(drop=True)

    POL = get_col_safe(df, 0)
    POD = get_col_safe(df, 1)

    Place = get_col_safe(df, 2)
    if Place is not None:
        Place = Place.astype(str).str.strip()

    RoutingNote = get_col_safe(df, 3)
    Carrier = get_col_safe(df, 5)
    Eff = get_col_safe(df, 6)
    Exp = get_col_safe(df, 7)
    Commodity = get_col_safe(df, 9) if rate_type == "FAK" else None
    Contract = get_col_safe(df, 11)

    # ALL-IN hiện tại (còn đang bao gồm PUC "sai" trong file FAK gốc)
    c20 = get_col_safe(df, 12)
    c40 = get_col_safe(df, 13)
    c40hq = get_col_safe(df, 14)
    c45 = get_col_safe(df, 15)
    c40nor = get_col_safe(df, 16)

    # PSS/PUC input (AM–AP: col 38–41)
    puc20_raw = get_col_safe(df, 38)  # 20GP
    puc40_raw = get_col_safe(df, 39)  # 40GP
    puc40hq_raw = get_col_safe(df, 40)  # 40HQ
    puc45_raw = get_col_safe(df, 41)  # 45HQ (KHÔNG DÙNG ĐỂ TRỪ)

    # Chuẩn hóa về số cho ALL-IN
    c20 = clean_amount_series(c20)
    c40 = clean_amount_series(c40)
    c40hq = clean_amount_series(c40hq)
    c45 = clean_amount_series(c45)
    c40nor = clean_amount_series(c40nor)

    # Chuẩn hóa về số cho PUC
    puc20 = clean_amount_series(puc20_raw)
    puc40 = clean_amount_series(puc40_raw)
    puc40hq = clean_amount_series(puc40hq_raw)
    puc45 = clean_amount_series(puc45_raw)  # vẫn đọc, nhưng KHÔNG trừ vào 45HQ

    price_df = pd.DataFrame(
        {
            "POL": POL,
            "POD": POD,
            "PlaceOfDelivery": Place,
            "RoutingNote": RoutingNote,
            "Carrier": Carrier,
            "EffectiveDate": Eff,
            "ExpirationDate": Exp,
            "ContractIdentifier": Contract,
            "CommodityType": Commodity,
            "20GP": c20,
            "40GP": c40,
            "40HQ": c40hq,
            "45HQ": c45,
            "40NOR": c40nor,
        }
    )

    # --------- TỰ ĐỘNG TRỪ PUC TRONG FAK_RAW (CMA/ONE/YML - SOC) ---------
    price_df["PUC20"] = puc20 if puc20 is not None else 0
    price_df["PUC40"] = puc40 if puc40 is not None else 0
    price_df["PUC40HQ"] = puc40hq if puc40hq is not None else 0
    price_df["PUC45"] = puc45 if puc45 is not None else 0  # KHÔNG TRỪ VÀO 45HQ

    mask_carrier = price_df["Carrier"].astype(str).str.upper().isin(
        [c.upper() for c in SOC_CARRIERS]
    )
    mask_soc_note = price_df["RoutingNote"].astype(str).str.upper().str.contains(
        "SOC", na=False
    )
    mask_apply_puc = mask_carrier & mask_soc_note

    if mask_apply_puc.any():
        for cont_col, puc_col in [
            ("20GP", "PUC20"),
            ("40GP", "PUC40"),
            ("40HQ", "PUC40HQ"),
        ]:
            if cont_col in price_df.columns:
                price_df.loc[mask_apply_puc, cont_col] = (
                    price_df.loc[mask_apply_puc, cont_col].fillna(0)
                    - price_df.loc[mask_apply_puc, puc_col].fillna(0)
                )

    price_df = price_df.drop(columns=["PUC20", "PUC40", "PUC40HQ", "PUC45"], errors="ignore")
    # ----------------------------------------------------------------------

    value_cols = ["20GP", "40GP", "40HQ", "45HQ", "40NOR"]
    value_cols = [c for c in value_cols if c in price_df.columns]

    df_out = (
        price_df.melt(
            id_vars=[
                "POL",
                "POD",
                "PlaceOfDelivery",
                "RoutingNote",
                "Carrier",
                "EffectiveDate",
                "ExpirationDate",
                "ContractIdentifier",
                "CommodityType",
            ],
            value_vars=value_cols,
            var_name="ContainerType",
            value_name="Amount",
        )
        .dropna(subset=["Amount"])
    )

    # ONE_SPECIAL RATE: gán CommodityType cố định
    if rate_type == "ONE_SPECIAL RATE":
        df_out["CommodityType"] = "FIX RATE"

    # Rút gọn / gom CommodityType cho ONE
    mask_one = df_out["Carrier"].astype(str).str.upper().str.contains("ONE", na=False)
    if mask_one.any():
        comm = df_out.loc[mask_one, "CommodityType"].astype(str)

        m_gar = comm.str.contains("GARMENT", case=False, na=False)
        comm.loc[m_gar] = "FAK: TPE1 - FAK Straight"

        m1 = comm.str.contains("FAK: TPE1 - FAK STRAIGHT", case=False, na=False)
        comm.loc[m1] = "FAK: TPE1 - FAK Straight"

        m2 = comm.str.contains("REEFER FAK", case=False, na=False)
        comm.loc[m2] = "REEFER FAK"

        m3 = comm.str.contains("SHORT TERM GDSM", case=False, na=False)
        comm.loc[m3] = "SHORT TERM GDSM"

        m4 = comm.str.contains("TPE9", case=False, na=False) & comm.str.contains(
            "GROUP SOC", case=False, na=False
        )
        comm.loc[m4] = "S1– TPE9 – Group SOC"

        df_out.loc[mask_one, "CommodityType"] = comm

    df_out["Amount"] = clean_amount_series(df_out["Amount"])
    df_out["RateType"] = rate_type
    df_out["SourceFile"] = source_file
    df_out = df_out.drop_duplicates()

    final_cols = [
        "POL",
        "POD",
        "PlaceOfDelivery",
        "RoutingNote",
        "Carrier",
        "EffectiveDate",
        "ExpirationDate",
        "ContractIdentifier",
        "ContainerType",
        "Amount",
        "CommodityType",
        "RateType",
        "SourceFile",
    ]
    return df_out[final_cols]


# ========= PARSER CHO HPL_SCFI =========
def parse_scfi(raw_df, source_file):
    """
    HPL_SCFI:
    A: POL
    B: POD
    C: PlaceOfDelivery
    D: Effective
    E: Expiration
    F–H: 20, 40, 40HC (map 20GP, 40GP, 40HQ)
    """
    if raw_df.shape[0] < 3:
        return pd.DataFrame()

    df = raw_df.iloc[2:].reset_index(drop=True)

    POL = get_col_safe(df, 0)
    POD = get_col_safe(df, 1)

    Place = get_col_safe(df, 2)
    if Place is not None:
        Place = Place.astype(str).str.strip()

    Eff = get_col_safe(df, 3)
    Exp = get_col_safe(df, 4)
    c20 = get_col_safe(df, 5)
    c40 = get_col_safe(df, 6)
    c40hq = get_col_safe(df, 7)

    price_df = pd.DataFrame(
        {
            "POL": POL,
            "POD": POD,
            "PlaceOfDelivery": Place,
            "RoutingNote": None,
            "Carrier": "HPL",
            "EffectiveDate": Eff,
            "ExpirationDate": Exp,
            "ContractIdentifier": None,
            "CommodityType": None,
            "20GP": c20,
            "40GP": c40,
            "40HQ": c40hq,
        }
    )

    df_out = (
        price_df.melt(
            id_vars=[
                "POL",
                "POD",
                "PlaceOfDelivery",
                "RoutingNote",
                "Carrier",
                "EffectiveDate",
                "ExpirationDate",
                "ContractIdentifier",
                "CommodityType",
            ],
            value_vars=["20GP", "40GP", "40HQ"],
            var_name="ContainerType",
            value_name="Amount",
        )
        .dropna(subset=["Amount"])
    )

    df_out["Amount"] = clean_amount_series(df_out["Amount"])
    df_out["RateType"] = "HPL_SCFI"
    df_out["SourceFile"] = source_file

    final_cols = [
        "POL",
        "POD",
        "PlaceOfDelivery",
        "RoutingNote",
        "Carrier",
        "EffectiveDate",
        "ExpirationDate",
        "ContractIdentifier",
        "ContainerType",
        "Amount",
        "CommodityType",
        "RateType",
        "SourceFile",
    ]
    return df_out[final_cols]


# ========= XÁC ĐỊNH RATE TYPE TỪ TÊN FILE =========
def detect_rate_type_from_name(fname: str) -> str | None:
    """Xác định RateType từ tên file, ONE_SPECIAL RATE thay thế ONE_FIX."""
    name_upper = fname.upper()

    if "SCFI" in name_upper:
        return "HPL_SCFI"
    if "FAK" in name_upper:
        return "FAK"

    if "ONE_SPECIAL RATE" in name_upper:
        return "ONE_SPECIAL RATE"
    if "ONE_SPECIAL" in name_upper and "RATE" in name_upper:
        return "ONE_SPECIAL RATE"

    if "ONE_FIX" in name_upper:
        return "ONE_SPECIAL RATE"
    if "ONE FIX" in name_upper:
        return "ONE_SPECIAL RATE"
    if "ONE-FIX" in name_upper:
        return "ONE_SPECIAL RATE"
    if "FIX" in name_upper and "ONE" in name_upper:
        return "ONE_SPECIAL RATE"
    if "FIX" in name_upper:
        return "ONE_SPECIAL RATE"

    return None


# ========= LOAD PUC =========
def load_puc():
    if not os.path.exists(puc_file):
        print(f"[!] Không tìm thấy file PUC_SOC.xlsx tại: {puc_file} -> bỏ qua bước PUC.")
        return None

    df = pd.read_excel(puc_file, sheet_name="PUC_SOC")
    df.columns = [str(c).strip() for c in df.columns]

    required = ["PlaceOfDelivery", "20DC", "40HC"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"PUC_SOC.xlsx thiếu cột: {missing}")

    for col in ["20DC", "40HC"]:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.strip()
            .replace({"": pd.NA, "TBA": pd.NA, "N/A": pd.NA})
            .astype(float)
        )

    df["CityKey"] = df["PlaceOfDelivery"].astype(str).str.upper().str.strip()
    return df


def build_city_key_for_master(pod: str, puc_cities_upper: list[str]) -> str:
    """
    Ưu tiên:
    1. Nếu POD chứa tên city PUC nào -> lấy city đó
    2. Ngược lại -> phần trước '(' rồi trước ',' làm city
    """
    if pd.isna(pod):
        return ""
    up = str(pod).upper()
    matches = [city for city in puc_cities_upper if city in up]
    if matches:
        return sorted(matches, key=len, reverse=True)[0]
    base = up.split("(")[0]
    base = base.split(",")[0].strip()
    return base


def apply_puc_to_df(df_master: pd.DataFrame) -> pd.DataFrame:
    """
    Cộng PUC trực tiếp vào Amount cho CMA/ONE/YML SOC.
    KHÔNG CỘNG PUC CHO 45HQ.
    """
    puc_df = load_puc()
    if puc_df is None:
        return df_master

    df = df_master.copy()
    df["ContainerNorm"] = normalize_container(df["ContainerType"])

    if "PlaceOfDelivery" not in df.columns:
        raise ValueError("Master không có cột PlaceOfDelivery")

    puc_cities_upper = puc_df["CityKey"].dropna().unique().tolist()
    df["CityKey"] = df["PlaceOfDelivery"].apply(
        lambda x: build_city_key_for_master(x, puc_cities_upper)
    )

    df = df.merge(
        puc_df[["CityKey", "20DC", "40HC"]],
        on="CityKey",
        how="left",
        suffixes=("", "_PUC"),
    )

    df["PUC_selected"] = np.where(
        df["ContainerNorm"] == "20GP",
        df["20DC"],
        np.where(df["ContainerNorm"].isin(["40GP", "40HQ"]), df["40HC"], np.nan),
    )

    mask_carrier = df["Carrier"].astype(str).str.upper().isin(
        [c.upper() for c in SOC_CARRIERS]
    )
    mask_soc_note = df["RoutingNote"].astype(str).str.upper().str.contains(
        "SOC", na=False
    )
    mask_apply = mask_carrier & mask_soc_note

    puc_series = df.loc[mask_apply, "PUC_selected"].fillna(0)
    df.loc[mask_apply, "Amount"] = df.loc[mask_apply, "Amount"] + puc_series

    df = df.drop(
        columns=[
            c
            for c in ["ContainerNorm", "CityKey", "20DC", "40HC", "PUC_selected"]
            if c in df.columns
        ]
    )

    return df


# ========= CHUẨN HÓA POD VÀ PLACEOFDELIVERY =========
def load_port_mapping():
    port_file = os.path.join(data_dir, "Port_Code_Mapping_Final.xlsx")
    if not os.path.exists(port_file):
        print(f"[!] Không tìm thấy file Port_Code_Mapping_Final.xlsx tại: {port_file}")
        return {}

    df = pd.read_excel(port_file)
    df.columns = [str(c).strip().upper() for c in df.columns]
    df = df.dropna(subset=["PORTNAME", "PORTCODE"])
    df["PORTNAME"] = df["PORTNAME"].astype(str).str.upper().str.strip()
    df["PORTCODE"] = df["PORTCODE"].astype(str).str.strip()
    return dict(zip(df["PORTNAME"], df["PORTCODE"]))


def normalize_pod_column(df, pod_map):
    df = df.copy()
    df["POD"] = df["POD"].apply(
        lambda x: pod_map.get(str(x).strip().upper(), str(x).strip())
        if pd.notna(x)
        else x
    )
    return df


def normalize_place_of_delivery_column(df):
    df = df.copy()
    df["PlaceOfDelivery"] = df["PlaceOfDelivery"].astype(str).str.upper().str.strip()
    return df


# ========= CHUẨN HÓA COMMODITY THEO HÃNG =========
def normalize_commodity(df_master: pd.DataFrame) -> pd.DataFrame:
    """
    Chuẩn hóa CommodityType theo từng hãng...
    """
    if "CommodityType" not in df_master.columns or "Carrier" not in df_master.columns:
        return df_master

    df = df_master.copy()

    carrier_upper = df["Carrier"].astype(str).str.upper()
    comm = df["CommodityType"].astype(str)

    # --- COSCO ---
    mask_cosco = carrier_upper.str.contains("COSCO", na=False)

    mask_cosco_fak_ex = mask_cosco & comm.str.contains(
        "FAK (Excluding Garment)", case=False, na=False, regex=False
    )
    df.loc[mask_cosco_fak_ex, "CommodityType"] = "FAK"

    mask_cosco_gar = mask_cosco & comm.str.contains(
        "Garments/Textile/Consol", case=False, na=False, regex=False
    )
    df.loc[mask_cosco_gar, "CommodityType"] = "GARMENT"

    # --- EMC ---
    mask_emc = carrier_upper.str.contains("EMC", na=False)
    mask_emc_rate1 = mask_emc & comm.str.contains(
        "RATE 1 - GENERAL CARGO", case=False, na=False, regex=False
    )
    df.loc[mask_emc_rate1, "CommodityType"] = "RATE 1"

    # --- HPL ---
    mask_hpl = carrier_upper.str.contains("HPL", na=False)

    mask_hpl_fak_inc = mask_hpl & comm.str.contains(
        "FAK INCLUDING GARMENT", case=False, na=False, regex=False
    )
    df.loc[mask_hpl_fak_inc, "CommodityType"] = "FAK"

    # --- YML ---
    mask_yml = carrier_upper.str.contains("YML", na=False)
    pattern_yml_fak = "FAK (NON-HAZ, EXCLUDING REEFER/ SHIPS/ BOATS/ VEHICLES/ CARS)"

    mask_yml_groupA = (
        mask_yml
        & comm.str.contains("GROUP A", case=False, na=False, regex=False)
        & comm.str.contains(pattern_yml_fak, case=False, na=False, regex=False)
    )
    df.loc[mask_yml_groupA, "CommodityType"] = "GROUP A"

    mask_yml_fak = (
        mask_yml
        & comm.str.contains(pattern_yml_fak, case=False, na=False, regex=False)
        & ~mask_yml_groupA
    )
    df.loc[mask_yml_fak, "CommodityType"] = "FAK"

    return df


# ========= XOAY DỌC -> NGANG =========
def make_horizontal_output(df_long: pd.DataFrame) -> pd.DataFrame:
    """
    Dọc (ContainerType/Amount) -> ngang (20GP, 40GP...).
    """
    if df_long.empty:
        return df_long.copy()

    df_long = df_long.copy()

    wanted_keys = [
        "POL",
        "POD",
        "PlaceOfDelivery",
        "RoutingNote",
        "Carrier",
        "EffectiveDate",
        "ExpirationDate",
        "ContractIdentifier",
        "CommodityType",
        "RateType",
    ]
    index_cols = [c for c in wanted_keys if c in df_long.columns]

    extra_keys = [
        c
        for c in df_long.columns
        if c not in index_cols and c not in ["ContainerType", "Amount"]
    ]
    index_cols = index_cols + extra_keys

    for col in index_cols:
        df_long[col] = df_long[col].fillna("")

    wide = (
        df_long.pivot_table(
            index=index_cols,
            columns="ContainerType",
            values="Amount",
            aggfunc="first",
        )
        .reset_index()
    )

    if wide.columns.name:
        wide.columns.name = None

    cont_order = ["20GP", "40GP", "40HQ", "45HQ", "40NOR"]
    key_cols = [c for c in wide.columns if c not in cont_order]
    other = [c for c in wide.columns if c not in key_cols and c not in cont_order]

    final_cols = key_cols + [c for c in cont_order if c in wide.columns] + other
    return wide[final_cols]


# ========= HÀM SET WIDTH CHO SHEET MASTER =========
def pixels_to_width(pixels: int) -> float:
    """
    Xấp xỉ convert pixel -> Excel column width (Calibri 11).
    """
    return max((pixels - 5) / 7.0, 0.0)


def set_master_column_widths(ws):
    """
    Set width cho các cột A..O trên sheet Master theo yêu cầu.
    """
    pixel_map = {
        "A": 80,
        "B": 141,
        "C": 141,
        "D": 125,
        "E": 85,
        "F": 112,
        "G": 112,
        "H": 146,
        "I": 150,
        "J": 80,
        "K": 80,
        "L": 80,
        "M": 80,
        "N": 80,
        "O": 80,
    }

    for col, px in pixel_map.items():
        ws.column_dimensions[col].width = pixels_to_width(px)


# ========= CHUẨN HÓA 1 FILE (KHÔNG GHI EXCEL TRUNG GIAN) =========
def normalize_file(filepath):
    fname = os.path.basename(filepath)
    print(f"\n[+] Đang xử lý RAW: {fname}")

    rate_type = detect_rate_type_from_name(fname)
    if rate_type is None:
        print("    -> Không xác định được RateType từ tên file, bỏ qua.")
        return pd.DataFrame()

    file_size = os.path.getsize(filepath) / (1024 * 1024)  # MB
    if file_size > 10:
        chunks = []
        for chunk in pd.read_excel(
            filepath,
            sheet_name=None,
            chunksize=5000,
            header=None,
            engine="openpyxl",
        ):
            chunks.append(chunk)
        raw_df = pd.concat(chunks)
    else:
        raw_df = read_excel_safe(filepath, header=None)

    if raw_df.empty:
        print("    -> File rỗng, bỏ qua.")
        return pd.DataFrame()

    if rate_type in ["FAK", "ONE_SPECIAL RATE"]:
        df_out = parse_fak_or_fix(raw_df, rate_type, fname)
    else:
        df_out = parse_scfi(raw_df, fname)

    if df_out.empty:
        print("    -> Không trích được dữ liệu hợp lệ.")
        return pd.DataFrame()

    print(f"    -> Đã chuẩn hóa (dạng dọc): {len(df_out)} dòng.")
    return df_out


# ========= GỘP MASTER + ÁP PUC VÀ LỌC EXPIRATION =========
def combine_all(list_df, include_expired: bool = False, cutoff_date: date | None = None):
    """
    Gộp toàn bộ df_long đã normalize và:
    - Nếu include_expired = False (mặc định):
        * Lọc bỏ mọi dòng có ExpirationDate < cutoff_date (hoặc hôm nay nếu cutoff_date=None)
        * Tức là 4-DEC thì 30-NOV, 2-DEC, 3-DEC đều bị loại.
    - Nếu include_expired = True:
        * Giữ FULL lịch sử, không cắt ExpirationDate.
    Sau đó: áp PUC, chuẩn hóa, xoay ngang và ghi Master_FullPricing.xlsx.
    """
    if not list_df:
        print("Không có dữ liệu hợp lệ để gộp!")
        return

    master = pd.concat(list_df, ignore_index=True)

    print("\n[THỐNG KÊ] Số dòng theo RateType (FULL, chưa cắt Exp):")
    print(master["RateType"].value_counts())

    filtered = master.copy()

    # ✅ Lọc ExpirationDate theo ngày
    if not include_expired:
        # Nếu không truyền cutoff_date thì dùng ngày hôm nay
        if cutoff_date is None:
            cutoff_date = datetime.today().date()

        # Parse ExpirationDate -> datetime.date
        exp_parsed = pd.to_datetime(filtered["ExpirationDate"], errors="coerce").dt.date

        # Giữ:
        #   - Dòng không có ExpirationDate (NaT) => coi như "no expiry"
        #   - Hoặc ExpirationDate >= cutoff_date
        mask = exp_parsed.isna() | (exp_parsed >= cutoff_date)
        before = len(filtered)
        filtered = filtered[mask].copy()
        after = len(filtered)
        print(
            f"[FILTER] Lọc ExpirationDate >= {cutoff_date} "
            f"(giữ cả dòng không có ExpirationDate): {before} -> {after} dòng."
        )
    else:
        print("[FILTER] include_expired=True -> GIỮ TOÀN BỘ lịch sử ExpirationDate (không lọc).")

    # Áp PUC (chuẩn) từ file PUC_SOC.xlsx
    filtered = apply_puc_to_df(filtered)

    # Chuẩn hóa ngày về dạng ngắn (sau filter)
    filtered["EffectiveDate"] = format_short_date(filtered["EffectiveDate"])
    filtered["ExpirationDate"] = format_short_date(filtered["ExpirationDate"])

    # Chuẩn hóa PlaceOfDelivery (bỏ space)
    filtered["PlaceOfDelivery"] = filtered["PlaceOfDelivery"].astype(str).str.strip()

    # ÁP DỤNG CHUẨN HÓA POD
    pod_map = load_port_mapping()
    filtered = normalize_pod_column(filtered, pod_map)

    # CHUẨN HÓA PlaceOfDelivery
    filtered = normalize_place_of_delivery_column(filtered)

    # CHUẨN HÓA COMMODITY THEO HÃNG
    filtered = normalize_commodity(filtered)

    # Bỏ SourceFile khỏi output chính
    filtered_no_src = filtered.drop(columns=["SourceFile"], errors="ignore")

    # MASTER PRICING: XOAY DỌC -> NGANG
    master_wide = make_horizontal_output(filtered_no_src)

    with pd.ExcelWriter(master_file, engine="openpyxl") as writer:
        # Sheet Master (ngang)
        master_wide.to_excel(writer, index=False, sheet_name="Master")
        ws = writer.sheets["Master"]

        # Tô màu header
        header_fill = PatternFill(
            start_color="CCFFCC",
            end_color="CCFFCC",
            fill_type="solid",
        )
        for cell in ws[1]:
            cell.fill = header_fill

        # Set width các cột A..O
        set_master_column_widths(ws)

        # Sheet Master_Long (dọc, sau khi lọc Exp + PUC + chuẩn commodity)
        filtered_no_src.to_excel(writer, index=False, sheet_name="Master_Long")

        # ✅ TẠO SHEET PHIÊN BẢN TỪ FILE RAW
        fak_files = [f for f in os.listdir(raw_dir) if "FAK" in f.upper()]
        if fak_files:
            latest_fak = sorted(fak_files)[-1]
            version_name = extract_version_from_filename(latest_fak)
            write_version_sheet(writer, version_name, filtered_no_src, fak_files)
            print(f"[VERSION] Đã thêm sheet version: {version_name}")
        else:
            print("[VERSION] Không tìm thấy file FAK trong Raw/ để tạo version.")

        # ✅ ĐÍNH KÈM SHEET LỊCH TÀU (Schedule)
        if os.path.exists(schedule_file):
            try:
                df_sched = pd.read_excel(schedule_file)
                df_sched.to_excel(writer, index=False, sheet_name="Schedule")
                print(f"[SCHEDULE] Đã đính kèm sheet 'Schedule' từ file: {schedule_file}")
            except Exception as e:
                print(f"[SCHEDULE] Lỗi khi đọc/ghi Schedule.xlsx: {e}")
        else:
            print(f"[SCHEDULE] Không tìm thấy file Schedule.xlsx tại: {schedule_file} -> bỏ qua.")

    print(f"\n[HOÀN TẤT] MASTER FILE: {master_file}")
    print(
        f"    -> Tổng số dòng Master_Long (sau filter Exp + PUC + chuẩn commodity): {len(filtered_no_src)}"
    )


# ========= HÀM CHO STREAMLIT GỌI (UPGRADE) =========
def normalize_all_from_streamlit(
    raw_dir_override=None,
    data_dir_override=None,
    include_expired: bool = False,
    cutoff_date: date | None = None,
):
    """
    Wrapper để chạy toàn bộ pipeline normalize (như main()) ngay trong Streamlit.

    Tham số mới:
    - include_expired: True -> giữ cả giá hết hạn để preview.
    - cutoff_date: nếu muốn fix theo một ngày cụ thể (vd 2025-12-04),
                   còn None -> dùng ngày hôm nay.
    """
    global raw_dir, data_dir, master_file, puc_file, schedule_file

    # Lưu lại cấu hình cũ
    old_raw_dir = raw_dir
    old_data_dir = data_dir
    old_master_file = master_file
    old_puc_file = puc_file
    old_schedule_file = schedule_file

    import streamlit as st
    import threading

    try:
        # Override nếu được truyền từ Streamlit
        if raw_dir_override is not None:
            raw_dir = str(raw_dir_override)
        if data_dir_override is not None:
            data_dir = str(data_dir_override)
            master_file = os.path.join(data_dir, "Master_FullPricing.xlsx")
            puc_file = os.path.join(data_dir, "PUC_SOC.xlsx")
            schedule_file = os.path.join(data_dir, "Schedule.xlsx")

        if not os.path.isdir(raw_dir):
            raise FileNotFoundError(f"Thư mục Raw không tồn tại: {raw_dir}")

        files = [f for f in os.listdir(raw_dir) if f.lower().endswith(".xlsx")]
        if not files:
            raise FileNotFoundError("Không tìm thấy file trong thư mục Raw.")

        placeholder = st.empty()
        progress_bar = st.progress(0.0)

        def run_normalize():
            all_normalized = []
            for i, f in enumerate(files):
                df_norm = normalize_file(os.path.join(raw_dir, f))
                if not df_norm.empty:
                    all_normalized.append(df_norm)
                progress_bar.progress((i + 1) / len(files))

            combine_all(all_normalized, include_expired=include_expired, cutoff_date=cutoff_date)
            placeholder.success(f"Hoàn tất! Master file: {master_file}")

        thread = threading.Thread(target=run_normalize)
        thread.start()
        st.info("Đang normalize... (background, có thể refresh page)")

        return master_file

    finally:
        # Khôi phục cấu hình global ban đầu
        raw_dir = old_raw_dir
        data_dir = old_data_dir
        master_file = old_master_file
        puc_file = old_puc_file
        schedule_file = old_schedule_file


# ========= MAIN =========
def main(include_expired: bool = False, cutoff_date: date | None = None):
    """
    Chạy normalize từ CLI.
    - include_expired=False: chỉ giữ giá còn hạn tại cutoff_date/hôm nay.
    - include_expired=True: build Master FULL lịch sử, dùng xem/so sánh.
    """
    if not os.path.isdir(raw_dir):
        print(f"Thư mục Raw không tồn tại: {raw_dir}")
        return

    files = [f for f in os.listdir(raw_dir) if f.lower().endswith(".xlsx")]
    if not files:
        print("Không tìm thấy file trong thư mục Raw.")
        return

    all_normalized = []
    for f in files:
        df_norm = normalize_file(os.path.join(raw_dir, f))
        if not df_norm.empty:
            all_normalized.append(df_norm)

    combine_all(all_normalized, include_expired=include_expired, cutoff_date=cutoff_date)


if __name__ == "__main__":
    # Mặc định: chỉ giữ giá còn hạn (ExpirationDate >= hôm nay)
    main()
