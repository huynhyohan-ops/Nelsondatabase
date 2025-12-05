import streamlit as st
import pandas as pd
from contextlib import nullcontext
from datetime import date, datetime, timedelta

from common.helpers import DATA_DIR  # thư mục Data dùng chung

import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============================================================
# CẤU HÌNH & HẰNG SỐ
# ============================================================

SHIPMENT_FILE = DATA_DIR / "Shipments.xlsx"

CONTAINER_TYPES = ["20GP", "40HQ", "20RF", "40RF", "45HQ", "40NOR"]

CONTAINER_TEU_MAP = {
    "20GP": 1,
    "20RF": 1,
    "40HQ": 2,
    "40RF": 2,
    "45HQ": 2,
    "40NOR": 2,
}

STATUS_OPTIONS = [
    "Submit",
    "Keep Space",
    "Confirmed",
    "Send SI",
    "Hbl Issue",
    "In Transit",
    "Arrival",
    "Delivered",
    "Cancelled",
]

CARRIER_OPTIONS = ["ONE", "CMA", "ZIM", "YML", "HPL", "MSK", "COSCO", "MSC", "WHL"]

BASE_COLUMNS = [
    "Customer",
    "Customer Type",
    "Routing",
    "BKG NO",
    "HBL NO",
    "ETD",
    "ETA",
    "Container Type",
    "Quantity",
    "Volume",
    "Status",
    "Selling Rate",
    "Buying Rate",
    "Profit",
    "SI",
    "CY",
    "Carrier",
    "HDL FEE carrier",
]

CUSTOMER_TYPE_OPTIONS = ["Direct", "Coload"]

# KPI THÁNG
KPI_VOLUME = 80.0      # TEU / month
KPI_PROFIT = 15500.0   # USD / month

# Màu & icon cho từng trạng thái giúp nhìn nhanh bảng nhiều cột
STATUS_BADGE_STYLE = {
    "Submit": {"color": "#0ea5e9", "emoji": "📝"},
    "Keep Space": {"color": "#fbbf24", "emoji": "⏳"},
    "Confirmed": {"color": "#22c55e", "emoji": "✅"},
    "Send SI": {"color": "#3b82f6", "emoji": "📤"},
    "Hbl Issue": {"color": "#a855f7", "emoji": "📄"},
    "In Transit": {"color": "#6366f1", "emoji": "🚢"},
    "Arrival": {"color": "#2dd4bf", "emoji": "🛬"},
    "Delivered": {"color": "#10b981", "emoji": "📦"},
    "Cancelled": {"color": "#ef4444", "emoji": "✖"},
}


# ============================================================
# THÁNG 11/2025 -> 11/2026
# ============================================================

def get_month_key_range():
    """Tạo list tháng từ 11/2025 đến 11/2026 dạng YYYY-MM."""
    months = []
    year = 2025
    month = 11
    while True:
        months.append(f"{year}-{month:02d}")
        if year == 2026 and month == 11:
            break
        month += 1
        if month == 13:
            month = 1
            year += 1
    return months


# ============================================================
# LOAD / SAVE
# ============================================================

def _empty_month_df() -> pd.DataFrame:
    """Tạo DataFrame trống với đúng dtype (ETD/ETA là datetime)."""
    data = {}
    for c in BASE_COLUMNS:
        if c in ["ETD", "ETA", "SI", "CY"]:
            data[c] = pd.Series(dtype="datetime64[ns]")
        else:
            data[c] = pd.Series(dtype="object")
    return pd.DataFrame(data)


def compute_df_signature(df: pd.DataFrame) -> str:
    """Tạo chữ ký đơn giản cho dataframe để báo thay đổi (tránh mất dữ liệu)."""
    try:
        normalized = df.copy()
        normalized = normalized.fillna("")
        return str(pd.util.hash_pandas_object(normalized, index=True).sum())
    except Exception:
        return ""


def load_month_df(month_key: str) -> pd.DataFrame:
    """
    Load data 1 tháng từ Shipments.xlsx.
    Nếu chưa có sheet, trả về DataFrame trống với cấu trúc chuẩn.
    """
    if SHIPMENT_FILE.exists():
        try:
            sheets = pd.read_excel(SHIPMENT_FILE, sheet_name=None)
            if month_key in sheets:
                df = sheets[month_key]

                # đảm bảo đủ cột
                for c in BASE_COLUMNS:
                    if c not in df.columns:
                        df[c] = None

                # ETD/ETA parse sang datetime
                for col in ["ETD", "ETA", "SI", "CY"]:
                    df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

                df = df[BASE_COLUMNS + [c for c in df.columns if c not in BASE_COLUMNS]]
                return df
        except Exception as e:
            st.error(f"Lỗi đọc file Shipments.xlsx: {e}")

    return _empty_month_df()


def _autofit_excel_columns(path):
    """Sau khi ghi file, tự động kéo rộng cột theo nội dung."""
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            # Tăng chiều cao header cho dễ đọc
            ws.row_dimensions[1].height = 20
            for col in ws.columns:
                max_length = 0
                column_index = col[0].column  # 1-based index
                for cell in col:
                    value = cell.value
                    if value is None:
                        continue
                    value_str = str(value)
                    if len(value_str) > max_length:
                        max_length = len(value_str)
                adjusted_width = max_length + 2
                col_letter = get_column_letter(column_index)
                ws.column_dimensions[col_letter].width = adjusted_width
        wb.save(path)
    except Exception as e:
        st.warning(f"Không auto-fit được cột trong Excel: {e}")


def save_month_df(month_key: str, df: pd.DataFrame):
    """
    Lưu DataFrame của 1 tháng vào file Shipments.xlsx.
    Giữ nguyên các sheet tháng khác + auto-fit cột.
    """
    SHIPMENT_FILE.parent.mkdir(parents=True, exist_ok=True)

    sheets = {}
    if SHIPMENT_FILE.exists():
        try:
            sheets = pd.read_excel(SHIPMENT_FILE, sheet_name=None)
        except Exception as e:
            st.error(f"Lỗi đọc file Shipments.xlsx khi lưu: {e}")
            sheets = {}

    sheets[month_key] = df

    with pd.ExcelWriter(SHIPMENT_FILE, engine="openpyxl") as writer:
        for name, sdf in sheets.items():
            sdf.to_excel(writer, sheet_name=name, index=False)

    # Auto-fit cột sau khi ghi file
    _autofit_excel_columns(SHIPMENT_FILE)


def load_all_shipments() -> pd.DataFrame:
    """Đọc toàn bộ file Shipments.xlsx và gắn nhãn tháng cho mục đích phân tích."""
    if not SHIPMENT_FILE.exists():
        return _empty_month_df()

    try:
        sheets = pd.read_excel(SHIPMENT_FILE, sheet_name=None)
    except Exception:
        return _empty_month_df()

    frames = []
    for name, sdf in sheets.items():
        df = sdf.copy()
        for col in BASE_COLUMNS:
            if col not in df.columns:
                df[col] = None

        for col in ["ETD", "ETA", "SI", "CY"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        df = compute_volume_profit(df)
        df["MonthKey"] = name
        frames.append(df)

    if not frames:
        return _empty_month_df()

    return pd.concat(frames, ignore_index=True)


# ============================================================
# TÍNH TOÁN VOLUME & PROFIT
# ============================================================

def compute_volume_profit(df: pd.DataFrame) -> pd.DataFrame:
    """Tính Volume (TEU) & Profit từ Container Type + Quantity + Selling/Buying."""
    df = df.copy()

    for c in ["Container Type", "Quantity", "Selling Rate", "Buying Rate"]:
        if c not in df.columns:
            df[c] = None

    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    df["Selling Rate"] = pd.to_numeric(df["Selling Rate"], errors="coerce").fillna(0.0)
    df["Buying Rate"] = pd.to_numeric(df["Buying Rate"], errors="coerce").fillna(0.0)

    # Volume (TEU) từng dòng
    volumes = []
    for _, row in df.iterrows():
        ctype = str(row.get("Container Type") or "").strip().upper()
        qty = row["Quantity"]
        teu_per_cont = CONTAINER_TEU_MAP.get(ctype, 0.0)
        volumes.append(qty * teu_per_cont)
    df["Volume"] = volumes

    # Profit từng dòng
    df["Profit"] = (df["Selling Rate"] - df["Buying Rate"]) * df["Quantity"]

    return df


def extract_destination(routing_value: str) -> str:
    """Lấy đích đến từ chuỗi routing (VD: "HPH - DENVER" -> "DENVER")."""
    if not isinstance(routing_value, str):
        return ""
    if "-" in routing_value:
        return routing_value.split("-")[-1].strip().upper()
    return routing_value.strip().upper()


# ============================================================
# ALERT SI / CY + KPI UTILS (dùng cho Dashboard)
# ============================================================

def find_alerts(df: pd.DataFrame, column_name: str, hours_before: int = 48) -> pd.DataFrame:
    """
    Lọc những lô có SI/CY trong vòng `hours_before` tới (so với thời điểm hiện tại).
    Cột SI/CY lưu dạng text, cố gắng parse thành datetime.
    """
    if column_name not in df.columns or df.empty:
        return pd.DataFrame(columns=df.columns)

    now = datetime.now()
    horizon = now + timedelta(hours=hours_before)

    parsed = pd.to_datetime(df[column_name], errors="coerce", dayfirst=True)
    mask = (parsed.notna()) & (parsed >= now) & (parsed <= horizon)

    alert_df = df[mask].copy()
    if not alert_df.empty:
        alert_df[column_name + "_parsed"] = parsed[mask]
    return alert_df


def find_eta_alerts(df: pd.DataFrame, days_before: int = 7) -> pd.DataFrame:
    """Lọc lô có ETA trong vòng `days_before` để nhắc thanh toán/thu tiền."""
    if "ETA" not in df.columns or df.empty:
        return pd.DataFrame(columns=df.columns)

    now = datetime.now()
    horizon = now + timedelta(days=days_before)
    parsed = pd.to_datetime(df["ETA"], errors="coerce")
    mask = (parsed.notna()) & (parsed >= now) & (parsed <= horizon)

    eta_df = df[mask].copy()
    if not eta_df.empty:
        eta_df["ETA_parsed"] = parsed[mask]
    return eta_df


def filter_by_timeframe(df: pd.DataFrame, timeframe: str) -> pd.DataFrame:
    """Lọc dataframe theo ETD trong khung thời gian (month/quarter/year)."""
    if "ETD" not in df.columns or df.empty:
        return df

    today = datetime.now().date()
    timeframe = timeframe.lower()

    if timeframe == "quarter":
        quarter = (today.month - 1) // 3 + 1
        start_month = 3 * (quarter - 1) + 1
        start_date = date(today.year, start_month, 1)
        end_month = start_month + 2
        last_day = (date(today.year, end_month, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        end_date = last_day
    elif timeframe == "year":
        start_date = date(today.year, 1, 1)
        end_date = date(today.year, 12, 31)
    else:
        start_date = date(today.year, today.month, 1)
        next_month = start_date + timedelta(days=32)
        end_date = date(next_month.year, next_month.month, 1) - timedelta(days=1)

    etd_parsed = pd.to_datetime(df["ETD"], errors="coerce").dt.date
    mask = (etd_parsed.notna()) & (etd_parsed >= start_date) & (etd_parsed <= end_date)
    return df[mask]


def aggregate_kpi_categories(df: pd.DataFrame) -> pd.DataFrame:
    """Tạo bảng KPI theo nhóm Revenue / Orders / Conversion rate."""
    if df.empty:
        return pd.DataFrame(
            {
                "Category": ["Revenue", "Orders", "Conversion Rate"],
                "Value": [0.0, 0, 0.0],
                "Note": ["", "", ""],
            }
        )

    df_numeric = df.copy()
    df_numeric["Quantity"] = pd.to_numeric(df_numeric.get("Quantity"), errors="coerce").fillna(0)
    df_numeric["Selling Rate"] = pd.to_numeric(df_numeric.get("Selling Rate"), errors="coerce").fillna(0.0)

    revenue = float((df_numeric["Selling Rate"] * df_numeric["Quantity"]).sum())
    orders = int(len(df_numeric))
    status_series = df_numeric.get("Status", pd.Series(dtype=str)).fillna("").str.lower()
    converted = status_series.isin(["confirmed", "send si", "hbl issue", "in transit", "arrival", "delivered"])
    conversion_rate = float((converted.sum() / orders) * 100) if orders > 0 else 0.0

    return pd.DataFrame(
        {
            "Category": ["Revenue", "Orders", "Conversion Rate"],
            "Value": [revenue, orders, conversion_rate],
            "Note": [
                "Tổng doanh thu (Selling Rate x Quantity)",
                "Tổng số lô hàng trong khung thời gian",
                "% lô đã chuyển từ Submit/Keep Space sang trạng thái thực hiện",
            ],
        }
    )


def render_filter_state():
    """Centralize dashboard filter state so all charts/tables stay in sync."""

    default_state = st.session_state.get(
        "follow_filter_state",
        {
            "customer_type": "All",
            "timeframe": "Month",
            "show_loss": True,
            "customer_query": "",
        },
    )

    filter_row = st.columns([1.2, 1.2, 1])
    type_options = ["All"] + CUSTOMER_TYPE_OPTIONS

    with filter_row[0]:
        customer_type = st.selectbox(
            "Customer Type",
            type_options,
            index=type_options.index(default_state.get("customer_type", "All")),
            key="follow_filter_customer_type",
        )
    with filter_row[1]:
        timeframe = st.radio(
            "Phạm vi thời gian",
            ["Month", "Quarter", "Year"],
            index=["Month", "Quarter", "Year"].index(default_state.get("timeframe", "Month")),
            horizontal=True,
            key="follow_filter_timeframe",
        )
    with filter_row[2]:
        show_loss = st.toggle(
            "Hiển thị khách loss",
            value=default_state.get("show_loss", True),
            help="Bật/tắt bảng khách đã hơn 3 tháng chưa ship.",
            key="follow_filter_loss",
        )

    search_row = st.columns([1, 1])
    with search_row[0]:
        customer_query = st.text_input(
            "Lọc theo khách hàng (áp dụng cho biểu đồ/routing)",
            value=default_state.get("customer_query", ""),
            placeholder="Nhập tên khách...",
            key="follow_filter_customer_query",
        )
    with search_row[1]:
        st.caption("Các bộ lọc áp dụng đồng nhất cho biểu đồ, bảng và cảnh báo.")

    filter_state = {
        "customer_type": customer_type,
        "timeframe": timeframe,
        "show_loss": show_loss,
        "customer_query": customer_query,
    }

    has_changed = filter_state != default_state
    st.session_state["follow_filter_state"] = filter_state

    return filter_state, has_changed


def render_status_legend():
    """Hiển thị legend màu/icon cho trạng thái để đọc bảng nhanh hơn."""
    chips = []
    for status, meta in STATUS_BADGE_STYLE.items():
        chips.append(
            f"<div class='status-chip' style='border-color:{meta['color']};color:{meta['color']}'>"
            f"{meta['emoji']} {status}</div>"
        )

    st.markdown(
        """
        <div class='legend-wrap'>
            <div class='legend-title'>Status legend</div>
            <div class='chip-row'>%s</div>
        </div>
        """
        % "".join(chips),
        unsafe_allow_html=True,
    )


# ============================================================
# UI CHÍNH – FOLLOW SHIPMENT (NHẬP LIỆU + KPI MONTH + ALERT)
# ============================================================

def render_follow_shipment_page():
    """Trang Follow Shipment – theo dõi lô hàng theo tháng."""
    st.markdown(
        "<div class='section-title'>Follow Shipment – Theo dõi lô hàng</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='section-sub'>Quản lý trạng thái các lô hàng theo từng tháng (11/2025 → 11/2026). Mỗi tháng là 1 sheet trong Shipments.xlsx.</div>",
        unsafe_allow_html=True,
    )

    col_head1, col_head2 = st.columns([2, 1])
    with col_head1:
        st.markdown(
            """
            <div class='ribbon'>
                <h4>Follow Shipment workspace</h4>
                <p>Nhập liệu nhanh đa cột, cập nhật KPI tức thì, có cảnh báo thay đổi để tránh mất dữ liệu.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with col_head2:
        render_status_legend()

    # ---------- CSS cho KPI CARD ----------
    kpi_css = """
    <style>
    .kpi-card {
        background: #ffffff;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(15, 23, 42, 0.06);
        padding: 14px 18px;
        margin-top: 4px;
        margin-bottom: 12px;
    }
    .kpi-title {
        font-size: 11px;
        font-weight: 600;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        color: #9ca3af;
        margin-bottom: 2px;
    }
    .kpi-value {
        font-size: 22px;
        font-weight: 700;
        color: #111827;
        margin-bottom: 2px;
    }
    .kpi-sub {
        font-size: 12px;
        color: #6b7280;
    }
    .kpi-bar {
        width: 100%;
        height: 6px;
        border-radius: 999px;
        background: #e5e7eb;
        margin-top: 10px;
    }
    .kpi-bar-fill {
        height: 100%;
        border-radius: 999px;
        background: linear-gradient(90deg, #22c55e, #16a34a);
    }
    .legend-wrap {
        background: #f8fafc;
        border: 1px dashed #e5e7eb;
        border-radius: 12px;
        padding: 10px 12px;
    }
    .legend-title {
        font-size: 12px;
        font-weight: 700;
        color: #475569;
        margin-bottom: 6px;
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }
    .chip-row {
        display: flex;
        flex-wrap: wrap;
        gap: 6px;
    }
    .status-chip {
        border: 1px solid #e5e7eb;
        background: #fff;
        border-radius: 999px;
        padding: 4px 10px;
        font-size: 12px;
        font-weight: 600;
        display: inline-flex;
        align-items: center;
        gap: 6px;
    }
    .ribbon {
        background: linear-gradient(90deg, #0ea5e9, #6366f1);
        color: #fff;
        padding: 12px 14px;
        border-radius: 12px;
        box-shadow: 0 10px 24px rgba(99, 102, 241, 0.18);
    }
    .ribbon h4 {
        margin: 0;
        font-size: 16px;
        font-weight: 700;
    }
    .ribbon p {
        margin: 2px 0 0;
        font-size: 13px;
        opacity: 0.9;
    }
    .pill { 
        display: inline-block;
        padding: 4px 10px;
        border-radius: 999px;
        font-size: 12px;
        font-weight: 600;
        background: rgba(255,255,255,0.18);
        border: 1px solid rgba(255,255,255,0.35);
        margin-left: 8px;
    }
    .guard {
        display: flex;
        align-items: center;
        gap: 8px;
        padding: 10px 12px;
        border-radius: 12px;
        border: 1px solid #e5e7eb;
        background: #f9fafb;
    }
    .guard strong { color: #111827; }
    </style>
    """
    st.markdown(kpi_css, unsafe_allow_html=True)

    # ---------- Chọn tháng ----------
    month_keys = get_month_key_range()
    today = date.today()
    current_key = f"{today.year}-{today.month:02d}"
    default_index = month_keys.index(current_key) if current_key in month_keys else 0

    month_choice = st.selectbox(
        "Chọn tháng cần xem / cập nhật",
        options=month_keys,
        index=default_index,
        help="Format: YYYY-MM. Mỗi tháng là 1 sheet riêng trong file Shipments.xlsx",
    )
    st.caption(f"📁 File lưu trữ: `{SHIPMENT_FILE}` – Sheet: `{month_choice}`")

    # ---------- DÙNG SESSION_STATE GIỮ DATAFRAME THÁNG ----------
    state_key = f"shipment_df_{month_choice}"
    sig_key = f"shipment_df_sig_{month_choice}"
    if state_key in st.session_state:
        df_month = st.session_state[state_key]
    else:
        df_month = load_month_df(month_choice)
        st.session_state[state_key] = df_month

    # Ép kiểu (làm một lần cho state hiện tại)
    for col in ["ETD", "ETA"]:
        if col in df_month.columns:
            df_month[col] = pd.to_datetime(df_month[col], errors="coerce")

    # Các cột numeric nhập tay (không đụng Volume/Profit vì sẽ tính tự động)
    numeric_input_cols = [
        "Quantity",
        "Selling Rate",
        "Buying Rate",
        "HDL FEE carrier",
    ]
    for col in numeric_input_cols:
        if col in df_month.columns:
            df_month[col] = pd.to_numeric(df_month[col], errors="coerce")

    # Tính Volume & Profit cho state hiện tại
    df_month = compute_volume_profit(df_month)
    st.session_state[state_key] = df_month

    if sig_key not in st.session_state:
        st.session_state[sig_key] = compute_df_signature(df_month)

    # ============================================================
    # CHẾ ĐỘ NHẬP LIỆU NHANH
    # ============================================================

    fast_mode = st.toggle(
        "⚡ Ưu tiên nhập liệu nhanh (hạn chế tính KPI/Chart nặng khi đang nhập)",
        value=True,
        help="Bật khi anh cần paste / chỉnh sửa nhiều ô cho mượt. Khi cần xem biểu đồ chi tiết thì sang tab KPI Dashboard.",
    )

    st.markdown("### 🛠️ Làn nhập liệu nhanh & bảo vệ dữ liệu")
    guard_col, kpi_hint_col = st.columns([2, 1])
    with guard_col:
        st.markdown(
            """
            <div class='guard'>
              <div>🛡️</div>
              <div><strong>Khóa cấu trúc cột</strong> – tên cột & thứ tự mặc định được giữ nguyên để tránh lỡ tay rename hoặc mất dữ liệu.
              Dữ liệu chỉnh sửa sẽ hiển thị cảnh báo <em>chưa lưu</em>.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with kpi_hint_col:
        st.info(
            "Tips: nhập đủ ETD/ETA & Status để KPI tháng chạy chính xác; SI/CY tự cảnh báo 48h sắp tới.",
            icon="💡",
        )

    column_config = {
        "ETD": st.column_config.DateColumn(
            "ETD (Date)",
            format="DD-MMM-YYYY",
        ),
        "ETA": st.column_config.DateColumn(
            "ETA (Date)",
            format="DD-MMM-YYYY",
        ),
        "Container Type": st.column_config.SelectboxColumn(
            "Container Type",
            options=CONTAINER_TYPES,
            help="Chọn loại cont. Volume sẽ được tính tự động theo TEU.",
        ),
        "Status": st.column_config.SelectboxColumn(
            "Status",
            options=STATUS_OPTIONS,
        ),
        "Carrier": st.column_config.SelectboxColumn(
            "Carrier",
            options=["--"] + CARRIER_OPTIONS,
        ),
        "Customer Type": st.column_config.SelectboxColumn(
            "Customer Type",
            options=CUSTOMER_TYPE_OPTIONS,
            help="Xác định loại khách: Direct hoặc Coload",
        ),
        "Quantity": st.column_config.NumberColumn(
            "Quantity",
            help="Số lượng cont",
            min_value=0,
            step=1,
        ),
        "Volume": st.column_config.NumberColumn(
            "Volume (TEU)",
            help="Tự động tính: Quantity x TEU theo Container Type",
            disabled=True,
            format="%.1f",
        ),
        "Selling Rate": st.column_config.NumberColumn(
            "Selling Rate (USD/cont)",
            format="%.2f",
        ),
        "Buying Rate": st.column_config.NumberColumn(
            "Buying Rate (USD/cont)",
            format="%.2f",
        ),
        "Profit": st.column_config.NumberColumn(
            "Profit (USD)",
            help="Tự động tính: (Selling - Buying) x Quantity",
            disabled=True,
            format="%.2f",
        ),
        "HDL FEE carrier": st.column_config.NumberColumn(
            "HDL FEE carrier (USD)",
            format="%.2f",
        ),
        "SI": st.column_config.DatetimeColumn(
            "SI (cut-off)",
            format="DD-MMM-YYYY HH:mm",
        ),
        "CY": st.column_config.DatetimeColumn(
            "CY (cut-off)",
            format="DD-MMM-YYYY HH:mm",
        ),
    }

    # ============================================================
    # FORM THÊM NHANH 1 DÒNG
    # ============================================================

    with st.form("quick_add_form", clear_on_submit=True):
        st.markdown("#### ➕ Thêm nhanh một lô hàng (tránh phải cuộn bảng dài)")
        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.0, 1.0])
        with c1:
            customer = st.text_input("Customer", placeholder="Nhập tên khách")
            routing = st.text_input("Routing", placeholder="VD: HCM - LAX")
            customer_type = st.selectbox(
                "Customer Type", options=CUSTOMER_TYPE_OPTIONS, help="Chọn nhanh Direct hoặc Coload"
            )
        with c2:
            etd_default = date.today() + timedelta(days=7)
            eta_default = etd_default + timedelta(days=20)
            etd_quick = st.date_input("ETD", value=etd_default)
            eta_quick = st.date_input("ETA", value=eta_default)
        with c3:
            cont_type = st.selectbox("Container Type", options=CONTAINER_TYPES)
            qty = st.number_input("Quantity", min_value=1, value=1, step=1)
            bkg_no_quick = st.text_input("BKG NO", placeholder="Tùy chọn")
        with c4:
            status_quick = st.selectbox("Status", options=STATUS_OPTIONS, index=STATUS_OPTIONS.index("Submit"))
            carrier_quick = st.selectbox("Carrier", options=["-- chọn --"] + CARRIER_OPTIONS, index=0)

        col_rate1, col_rate2 = st.columns(2)
        with col_rate1:
            selling_quick = st.number_input("Selling Rate (USD/cont)", min_value=0.0, value=0.0, step=10.0)
            si_date = st.date_input("SI Date", value=etd_default, help="Dùng để cảnh báo SI gần đến hạn")
            si_time = st.time_input("SI Time", value=datetime.now().time().replace(minute=0, second=0, microsecond=0))
        with col_rate2:
            buying_quick = st.number_input("Buying Rate (USD/cont)", min_value=0.0, value=0.0, step=10.0)
            cy_date = st.date_input("CY Date", value=etd_default, help="Dùng để cảnh báo CY gần đến hạn")
            cy_time = st.time_input("CY Time", value=datetime.now().time().replace(minute=0, second=0, microsecond=0))

        submitted = st.form_submit_button("Thêm vào bảng", type="primary")

        if submitted:
            si_dt = datetime.combine(si_date, si_time) if si_date else None
            cy_dt = datetime.combine(cy_date, cy_time) if cy_date else None
            new_row = {
                "Customer": customer.strip() if customer else None,
                "Customer Type": customer_type,
                "Routing": routing.strip() if routing else None,
                "BKG NO": bkg_no_quick.strip() if bkg_no_quick else None,
                "HBL NO": None,
                "ETD": pd.to_datetime(etd_quick),
                "ETA": pd.to_datetime(eta_quick),
                "Container Type": cont_type,
                "Quantity": qty,
                "Volume": None,
                "Status": status_quick,
                "Selling Rate": selling_quick,
                "Buying Rate": buying_quick,
                "Profit": None,
                "SI": si_dt,
                "CY": cy_dt,
                "Carrier": None if carrier_quick == "-- chọn --" else carrier_quick,
                "HDL FEE carrier": None,
            }
            df_month = pd.concat([df_month, pd.DataFrame([new_row])], ignore_index=True)
            st.session_state[state_key] = df_month
            st.success("Đã thêm lô hàng vào bảng nhập liệu, tiếp tục chỉnh sửa nếu cần. SI/CY đã gắn giờ để cảnh báo chuẩn.")

    # ============================================================
    # TÌM KIẾM & CẬP NHẬT TRẠNG THÁI/THỜI GIAN NHANH
    # ============================================================

    st.markdown("### 🔎 Tìm shipment & chỉnh sửa nhanh")
    search_query = st.text_input(
        "Tìm theo Customer, Routing, BKG, HBL hoặc Carrier",
        placeholder="Nhập từ khóa để lọc các dòng cần cập nhật",
    )

    if search_query:
        mask = pd.Series(False, index=df_month.index)
        for col in ["Customer", "Routing", "BKG NO", "HBL NO", "Carrier", "Status"]:
            if col in df_month.columns:
                mask = mask | df_month[col].fillna("").astype(str).str.lower().str.contains(search_query.lower())

        filtered = df_month[mask].copy()
        if filtered.empty:
            st.info("Không tìm thấy shipment khớp từ khóa.")
        else:
            st.caption("Chỉnh trực tiếp các cột trạng thái, SI/ETD/ETA/CY hoặc Carrier & Customer Type.")

            # FIX kiểu dữ liệu cho các cột selectbox
            for col in ["Customer Type", "Container Type", "Status", "Carrier"]:
                if col in filtered.columns:
                    filtered[col] = filtered[col].apply(
                        lambda x: x[0] if isinstance(x, (list, tuple)) else x
                    )
                    filtered[col] = filtered[col].astype(str)

            quick_columns = [
                c
                for c in [
                    "Customer",
                    "Customer Type",
                    "Routing",
                    "BKG NO",
                    "HBL NO",
                    "Status",
                    "Carrier",
                    "SI",
                    "ETD",
                    "ETA",
                    "CY",
                ]
                if c in filtered.columns
            ]

            edited_search = st.data_editor(
                filtered[quick_columns],
                num_rows="dynamic",
                use_container_width=True,
                key=f"search_editor_{month_choice}",
                column_config=column_config,
            )

            if st.button("Áp dụng thay đổi vào bảng chính", type="primary"):
                # giữ index gốc để update
                edited_search.index = filtered.index
                df_month.update(edited_search)
                df_month = compute_volume_profit(df_month)
                st.session_state[state_key] = df_month
                st.success("Đã cập nhật các dòng được tìm thấy.")

    # ============================================================
    # BẢNG SHIPMENT (EDIT TRỰC TIẾP)
    # ============================================================

    st.markdown("### 📋 Bảng shipment của tháng đã chọn")
    st.caption("Cột cố định, không reorder/rename để tránh sai cấu trúc. Tăng tốc nhập bằng double click/paste multi-cell.")

    # FIX kiểu dữ liệu cho các cột selectbox trong df_month
    for col in ["Customer Type", "Container Type", "Status", "Carrier"]:
        if col in df_month.columns:
            df_month[col] = df_month[col].apply(
                lambda x: x[0] if isinstance(x, (list, tuple)) else x
            )
            df_month[col] = df_month[col].astype(str)

    extra_columns = [c for c in df_month.columns if c not in BASE_COLUMNS]
    column_order = BASE_COLUMNS + extra_columns

    edited = st.data_editor(
        df_month,
        use_container_width=True,
        num_rows="dynamic",
        key=f"shipment_editor_{month_choice}",
        column_config=column_config,
        column_order=column_order,
        hide_index=True,
    )

    # Tính lại Volume/Profit cho toàn bộ tháng sau khi edit & lưu vào session
    edited = compute_volume_profit(edited)
    st.session_state[state_key] = edited
    df_month = edited

    current_sig = compute_df_signature(df_month)
    saved_sig = st.session_state.get(sig_key)
    dirty = saved_sig != current_sig

    if dirty:
        st.warning(
            "⚠️ Bảng đã thay đổi nhưng chưa lưu. Nhấn Lưu tháng này để tránh mất dữ liệu khi đổi tháng hoặc reload.",
            icon="🛟",
        )
    else:
        st.success("✅ Bảng đã đồng bộ với dữ liệu lưu gần nhất.")

    # ============================================================
    # KPI MONTH + NÚT LƯU
    # ============================================================

    if "Status" in df_month.columns:
        status_clean = df_month["Status"].fillna("").astype(str).str.strip()
        mask_real = ~status_clean.isin(["Keep Space", "Cancelled"])
    else:
        mask_real = pd.Series(True, index=df_month.index)

    total_teu = df_month.loc[mask_real, "Volume"].sum() if "Volume" in df_month.columns else 0.0
    total_profit = df_month.loc[mask_real, "Profit"].sum() if "Profit" in df_month.columns else 0.0

    st.markdown("### KPI MONTH")

    vol_pct = total_teu / KPI_VOLUME if KPI_VOLUME > 0 else 0
    prof_pct = total_profit / KPI_PROFIT if KPI_PROFIT > 0 else 0

    vol_percent = vol_pct * 100
    prof_percent = prof_pct * 100

    col_v, col_p = st.columns(2)

    with col_v:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-title">VOLUME KPI</div>
              <div class="kpi-value">{total_teu:.1f} TEU ({vol_percent:.1f}%)</div>
              <div class="kpi-sub">Target: {KPI_VOLUME:.1f} TEU / month</div>
              <div class="kpi-bar">
                <div class="kpi-bar-fill" style="width: {min(vol_percent, 100):.1f}%"></div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_p:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-title">PROFIT KPI</div>
              <div class="kpi-value">{total_profit:,.0f} USD ({prof_percent:.1f}%)</div>
              <div class="kpi-sub">Target: {KPI_PROFIT:,.0f} USD / month</div>
              <div class="kpi-bar">
                <div class="kpi-bar-fill" style="width: {min(prof_percent, 100):.1f}%"></div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("### 💾 Lưu dữ liệu tháng này")
    col_save1, col_save2 = st.columns([1, 3])

    with col_save1:
        if st.button("💾 Lưu tháng này", type="primary", key=f"save_{month_choice}"):
            df_to_save = st.session_state[state_key].copy()
            df_to_save = df_to_save[df_to_save.notna().any(axis=1)]
            save_month_df(month_choice, df_to_save)
            st.session_state[sig_key] = compute_df_signature(df_to_save)
            st.success(f"Đã lưu dữ liệu cho tháng {month_choice} vào {SHIPMENT_FILE}")
            st.balloons()

    with col_save2:
        st.info("KPI chỉ tính các lô không phải Keep Space / Cancelled. Biểu đồ chi tiết xem ở tab KPI Dashboard.")

    # ============================================================
    # CẢNH BÁO SI / CY
    # ============================================================

    st.markdown("---")
    st.markdown("### ⏰ Cảnh báo SI / CY trong 48 giờ tới")

    si_alerts = find_alerts(df_month, "SI", hours_before=48)
    cy_alerts = find_alerts(df_month, "CY", hours_before=48)

    if si_alerts.empty and cy_alerts.empty:
        st.success("Hiện không có lô hàng nào có SI/CY trong vòng 48 giờ tới.")
    else:
        if not si_alerts.empty:
            st.error("🚨 SI trong 48 giờ:")
            show_cols = [
                c
                for c in [
                    "Customer",
                    "Routing",
                    "BKG NO",
                    "HBL NO",
                    "Carrier",
                    "Container Type",
                    "Quantity",
                    "SI",
                    "SI_parsed",
                ]
                if c in si_alerts.columns
            ]
            st.dataframe(si_alerts[show_cols], use_container_width=True)

        if not cy_alerts.empty:
            st.warning("⚠ CY trong 48 giờ:")
            show_cols = [
                c
                for c in [
                    "Customer",
                    "Routing",
                    "BKG NO",
                    "HBL NO",
                    "Carrier",
                    "Container Type",
                    "Quantity",
                    "CY",
                    "CY_parsed",
                ]
                if c in cy_alerts.columns
            ]
            st.dataframe(cy_alerts[show_cols], use_container_width=True)

    # ETA payment reminder
    st.markdown("### 🚨 Cảnh báo ETA 7 ngày tới (nhắc thanh toán/thu tiền)")
    eta_alerts = find_eta_alerts(df_month, days_before=7)
    if eta_alerts.empty:
        st.info("Không có lô ETA trong 7 ngày tới.")
    else:
        show_cols_eta = [
            c
            for c in [
                "Customer",
                "Routing",
                "BKG NO",
                "Status",
                "ETA",
                "ETA_parsed",
                "Selling Rate",
                "Quantity",
                "Profit",
            ]
            if c in eta_alerts.columns
        ]
        st.warning("Nhắc thu tiền / thanh toán cho các lô sắp đến:")
        st.dataframe(eta_alerts[show_cols_eta], use_container_width=True)
