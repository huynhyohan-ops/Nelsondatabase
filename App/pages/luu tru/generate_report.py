from pathlib import Path
from datetime import datetime, timedelta, date

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===================== PATH CONFIG =====================

BASE_DIR = Path(__file__).resolve().parents[1]  # PricingSystem/
DATA_DIR = BASE_DIR / "Data"
OUTPUT_DIR = BASE_DIR / "Output" / "Reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SHIPMENTS_FILE = DATA_DIR / "Shipments.xlsx"


# ===================== ISO WEEK HELPERS =====================

def iso_week_range(year: int, week: int):
    """Trả về (monday, sunday) của tuần ISO (giống logic Apps Script)."""
    jan4 = date(year, 1, 4)
    jan4_weekday = jan4.isoweekday()  # 1=Mon..7=Sun
    monday = jan4 + timedelta(days=(week - 1) * 7 - (jan4_weekday - 1))
    sunday = monday + timedelta(days=6)
    return monday, sunday


def get_current_iso_week_and_year():
    today = datetime.today().date()
    iso_year, iso_week, _ = today.isocalendar()
    return iso_year, iso_week


# ===================== LOAD SHIPMENTS & FILTER BY WEEK =====================

def load_shipments_for_week(week: int, year: int):
    """
    Đọc tất cả sheet trong Shipments.xlsx,
    lọc các dòng có ETD thuộc tuần ISO (week, year).
    """
    if not SHIPMENTS_FILE.exists():
        raise FileNotFoundError(f"Không tìm thấy file shipments: {SHIPMENTS_FILE}")

    sheets = pd.read_excel(SHIPMENTS_FILE, sheet_name=None)
    records = []

    for sheet_name, df in sheets.items():
        if df.empty:
            continue

        for col in ["Customer", "Routing", "ETD", "Volume", "Profit"]:
            if col not in df.columns:
                df[col] = None

        etd_parsed = pd.to_datetime(df["ETD"], errors="coerce", dayfirst=True)

        for idx, row in df.iterrows():
            etd = etd_parsed.iloc[idx]
            if pd.isna(etd):
                continue

            iso_y, iso_w, _ = etd.isocalendar()
            if iso_y == year and iso_w == week:
                records.append(
                    {
                        "etd": etd,
                        "customer": row.get("Customer"),
                        "routing": row.get("Routing"),
                        "volume": float(row.get("Volume") or 0),
                        "profit": float(row.get("Profit") or 0),
                    }
                )

    records.sort(key=lambda r: r["etd"])
    return records


# ===================== TẠO TEMPLATE ĐẦU (HEADER + I + II.1) =====================

def create_report_workbook(week: int, year: int, monday: date, sunday: date):
    """
    Tạo workbook mới theo template báo cáo tuần:
    - Header
    - I/ TARGET THEO THÁNG
    - II/ title + II.1 (header, chưa dựng II.2/III/IV)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"TUAN_{week}"

    # Styles
    bold = Font(bold=True)
    header_font = Font(bold=True, size=18, color="0B5394")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    green_header_fill = PatternFill("solid", fgColor="93C47D")
    light_green_fill = PatternFill("solid", fgColor="D9EAD3")
    light_yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_medium = Border(top=medium, left=medium, right=medium, bottom=medium)

    # Column widths
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    # ===== HEADER =====
    ws["B1"] = "BÁO CÁO TUẦN"
    ws["B1"].font = header_font
    ws["B1"].alignment = left

    ws["E1"] = f"TUẦN: {week}"
    ws["E1"].font = bold

    ws["G1"] = "SALE'S NAME"
    ws["G1"].font = bold
    ws["H1"] = "Nelson"

    ws["B2"] = f"(Từ {monday.strftime('%d/%m')} đến {sunday.strftime('%d/%m')})"
    ws["B2"].font = Font(italic=True, color="6D6D6D")

    # ===== I/ TARGET THEO THÁNG =====
    ws.merge_cells("B4:H4")
    ws["B4"] = "I/ TARGET THEO THÁNG"
    ws["B4"].font = Font(bold=True, color="ED1C24")
    ws["B4"].alignment = center

    hdr1 = ["TUẦN", "SỐ CUỘC GỌI", "GẶP KH", "BÁO GIÁ", "KH MỚI", "PROFIT", "VOLUME"]
    start_row = 6
    for col, val in enumerate(hdr1, start=2):  # B..H
        cell = ws.cell(row=start_row, column=col, value=val)
        cell.font = bold
        cell.fill = light_green_fill
        cell.alignment = center
        cell.border = border_thin

    # 4 tuần: w, w+1, w+2, w+3
    weeks = [week + i for i in range(4)]
    data_start = start_row + 1  # row 7
    for i, w in enumerate(weeks):
        r = data_start + i
        ws.cell(row=r, column=2, value=w)
        for c in range(2, 2 + len(hdr1)):
            ws.cell(row=r, column=c).border = border_thin

    total_row = data_start + len(weeks)  # row 11
    ws.cell(row=total_row, column=2, value="TỔNG").font = bold
    for c in range(2, 2 + len(hdr1)):
        cell = ws.cell(row=total_row, column=c)
        cell.border = border_medium
        cell.fill = light_yellow_fill

    profit_col = 2 + 5  # G
    volume_col = 2 + 6  # H
    ws.cell(
        row=total_row,
        column=profit_col,
        value=f"=SUM({get_column_letter(profit_col)}{data_start}:{get_column_letter(profit_col)}{total_row-1})",
    )
    ws.cell(
        row=total_row,
        column=volume_col,
        value=f"=SUM({get_column_letter(volume_col)}{data_start}:{get_column_letter(volume_col)}{total_row-1})",
    )

    # ===== II/ BÁO CÁO CHI TIẾT CÔNG VIỆC SALE TRONG TUẦN =====
    sec2_title_row = total_row + 2
    ws.merge_cells(start_row=sec2_title_row, start_column=2, end_row=sec2_title_row, end_column=8)
    cell = ws.cell(row=sec2_title_row, column=2, value="II/ BÁO CÁO CHI TIẾT CÔNG VIỆC SALE TRONG TUẦN")
    cell.font = Font(bold=True, color="ED1C24")
    cell.alignment = center

    # ===== II.1 KHÁCH HÀNG ĐÃ SỬ DỤNG DỊCH VỤ TRONG TUẦN =====
    ii1_title_row = sec2_title_row + 1
    ws.merge_cells(start_row=ii1_title_row, start_column=2, end_row=ii1_title_row, end_column=6)
    cell = ws.cell(row=ii1_title_row, column=2, value="1/ KHÁCH HÀNG ĐÃ SỬ DỤNG DỊCH VỤ TRONG TUẦN")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = green_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ii1_header_row = ii1_title_row + 1
    hdr21 = ["STT", "ACCOUNT", "ROUTING", "VOLUME", "PROFIT"]
    for col, val in enumerate(hdr21, start=2):
        c = ws.cell(row=ii1_header_row, column=col, value=val)
        c.font = bold
        c.fill = light_green_fill
        c.alignment = center
        c.border = border_thin

    ii1_data_start = ii1_header_row + 1

    return wb, ws, {
        "tableI_first_row": data_start,
        "tableI_profit_col": profit_col,
        "tableI_volume_col": volume_col,
        "ii1_header_row": ii1_header_row,
        "ii1_data_start": ii1_data_start,
    }


# =============== HÀM DỰNG II.2 + III + IV SAU KHI BIẾT VỊ TRÍ ===============

def build_ii2_iii_iv(ws, start_row_ii2: int):
    """
    Xây II.2, III, IV bắt đầu từ dòng start_row_ii2.
    Giữ layout giống file gốc, nhưng khoảng cách phụ thuộc số shipment.
    """
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    green_header_fill = PatternFill("solid", fgColor="93C47D")
    light_green_fill = PatternFill("solid", fgColor="D9EAD3")
    body_fill = PatternFill("solid", fgColor="F9FFF9")
    light_yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    thin = Side(border_style="thin", color="000000")
    border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ===== II.2 BÁO GIÁ KHÁCH HÀNG =====
    ii2_title_row = start_row_ii2
    ws.merge_cells(start_row=ii2_title_row, start_column=2, end_row=ii2_title_row, end_column=7)
    cell = ws.cell(row=ii2_title_row, column=2, value="2/ BÁO GIÁ KHÁCH HÀNG")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = green_header_fill
    cell.alignment = center

    ii2_header_row = ii2_title_row + 1
    hdr22 = [
        "STT",
        "TÊN KHÁCH HÀNG",
        "CHI TIẾT TUYẾN BÁO GIÁ",
        "FEEDBACK",
        "NOTE",
        "SỐ LẦN BG/TUẦN",
    ]
    for col, val in enumerate(hdr22, start=2):  # B..G
        c = ws.cell(row=ii2_header_row, column=col, value=val)
        c.font = bold
        c.fill = light_green_fill
        c.alignment = center
        c.border = border_thin

    ii2_data_start = ii2_header_row + 1
    # 14 dòng data trống
    for r in range(ii2_data_start, ii2_data_start + 14):
        for c_idx in range(2, 2 + len(hdr22)):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = body_fill
            cell.border = border_thin

    ii2_total_row = ii2_data_start + 14
    ws.cell(row=ii2_total_row, column=2, value="TỔNG").font = bold
    ws.cell(row=ii2_total_row, column=2).fill = light_yellow_fill
    ws.cell(row=ii2_total_row, column=2).border = border_thin

    # SUM số lần báo giá (col G)
    g_col = 2 + len(hdr22) - 1
    ws.cell(
        row=ii2_total_row,
        column=g_col,
        value=f"=SUM({get_column_letter(g_col)}{ii2_data_start}:{get_column_letter(g_col)}{ii2_total_row-1})",
    )
    cell = ws.cell(row=ii2_total_row, column=g_col)
    cell.font = bold
    cell.fill = light_yellow_fill
    cell.border = border_thin

    # ===== III. ĐI GẶP KHÁCH HÀNG =====
    r3_title = ii2_total_row + 2
    ws.merge_cells(start_row=r3_title, start_column=2, end_row=r3_title, end_column=7)
    cell = ws.cell(row=r3_title, column=2, value="3/ ĐI GẶP KHÁCH HÀNG")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = green_header_fill
    cell.alignment = center

    r3_header = r3_title + 1
    hdr23 = ["STT", "TÊN KHÁCH HÀNG", "ĐỊA CHỈ", "THÔNG TIN", "ĐÁNH GIÁ", "KẾ HOẠCH"]
    for col, val in enumerate(hdr23, start=2):
        c = ws.cell(row=r3_header, column=col, value=val)
        c.font = bold
        c.fill = light_green_fill
        c.alignment = center
        c.border = border_thin

    r3_data_start = r3_header + 1
    for r in range(r3_data_start, r3_data_start + 8):
        for c_idx in range(2, 2 + len(hdr23)):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = body_fill
            cell.border = border_thin

    # ===== IV. KH MỚI =====
    r4_title = r3_data_start + 10  # 8 dòng + 2 dòng trống
    ws.merge_cells(start_row=r4_title, start_column=2, end_row=r4_title, end_column=7)
    cell = ws.cell(row=r4_title, column=2, value="4/ KH MỚI")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = green_header_fill
    cell.alignment = center

    r4_header = r4_title + 1
    hdr24 = ["STT", "TÊN KHÁCH HÀNG", "ĐỊA CHỈ", "DỊCH VỤ", "TIỀM NĂNG", "KẾ HOẠCH"]
    for col, val in enumerate(hdr24, start=2):
        c = ws.cell(row=r4_header, column=col, value=val)
        c.font = bold
        c.fill = light_green_fill
        c.alignment = center
        c.border = border_thin

    r4_data_start = r4_header + 1
    for r in range(r4_data_start, r4_data_start + 8):
        for c_idx in range(2, 2 + len(hdr24)):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = body_fill
            cell.border = border_thin


# ===================== FILL DATA VÀ GỌI BUILD II.2/III/IV =====================

def fill_report(week: int, year: int):
    monday, sunday = iso_week_range(year, week)
    shipments = load_shipments_for_week(week, year)

    report_path = OUTPUT_DIR / f"Weekly_Report_{year}_W{week:02d}.xlsx"
    if report_path.exists():
        report_path.unlink()

    wb, ws, anchors = create_report_workbook(week, year, monday, sunday)

    thin = Side(border_style="thin", color="000000")
    border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
    body_fill = PatternFill("solid", fgColor="F9FFF9")
    light_yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)

    # ----- II.1: fill shipments -----
    ii1_data_start = anchors["ii1_data_start"]
    total_profit = 0.0
    total_volume = 0.0

    for i, s in enumerate(shipments):
        r = ii1_data_start + i
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value=s["customer"])
        ws.cell(row=r, column=4, value=s["routing"])
        ws.cell(row=r, column=5, value=s["volume"])
        ws.cell(row=r, column=6, value=s["profit"])

        total_profit += s["profit"]
        total_volume += s["volume"]

        for c in range(2, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = border_thin
            cell.fill = body_fill

    total_row = ii1_data_start + len(shipments)
    ws.cell(row=total_row, column=2, value="TOTAL").font = bold
    for c in range(2, 7):
        cell = ws.cell(row=total_row, column=c)
        cell.border = border_thin
        if c in (5, 6):
            cell.fill = light_yellow_fill

    if shipments:
        ws.cell(row=total_row, column=5, value=total_volume)
        ws.cell(row=total_row, column=6, value=total_profit)

    # ----- Bảng I: cập nhật profit & volume -----
    tableI_first_row = anchors["tableI_first_row"]
    profit_col = anchors["tableI_profit_col"]
    volume_col = anchors["tableI_volume_col"]

    for offset in range(4):
        r = tableI_first_row + offset
        week_val = ws.cell(row=r, column=2).value
        if week_val == week:
            ws.cell(row=r, column=profit_col, value=total_profit)
            ws.cell(row=r, column=volume_col, value=total_volume)
            break

    # ----- Sau khi biết total_row của II.1, dựng II.2 + III + IV ngay phía dưới -----
    start_row_ii2 = total_row + 2  # 1 dòng TOTAL + 1 dòng trắng
    build_ii2_iii_iv(ws, start_row_ii2)

    wb.save(report_path)
    print(f"Đã tạo báo cáo tuần: {report_path}")
    print(f"Số shipment tuần {week}/{year}: {len(shipments)}")
    print(f"Tổng Volume: {total_volume}, Tổng Profit: {total_profit}"

    ws[f"D{i + 1}"] = f"=SUM(D5:D{i})"
    ws[f"E{i + 1}"] = f"=SUM(E5:E{i})"
    ws[f"B{i + 1}"] = "TOTAL"
    ws[f"B{i + 1}"].font = bold

    # Lưu file
    wb.save(report_path)

    # Trả về đường dẫn file
    return report_path

# ===================== ENTRY POINT =====================

if __name__ == "__main__":
    iso_year, iso_week = get_current_iso_week_and_year()
    # Test một tuần cụ thể nếu muốn:
    # iso_year, iso_week = 2025, 49

    fill_report(iso_week, iso_year)
